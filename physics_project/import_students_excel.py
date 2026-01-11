import mysql.connector
from openpyxl import load_workbook

# 只需要修改这里的文件名即可（支持 .xlsx）
EXCEL_PATH = 'students.xlsx'

# 数据库连接配置
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '123456',
    'database': 'physics_new3'
}

# 支持的表头字段（任选其一即可）
STUDENT_ID_HEADERS = {'学号', 'student_id', 'studentid', 'id'}


def find_student_id_column(header_cells):
    for idx, cell in enumerate(header_cells):
        value = str(cell.value).strip() if cell.value is not None else ''
        if value.lower() in STUDENT_ID_HEADERS:
            return idx
    return None


def main():
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    try:
        workbook = load_workbook(EXCEL_PATH)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            print('⚠️ Excel 文件为空')
            return

        header = rows[0]
        student_id_index = find_student_id_column(header)
        if student_id_index is None:
            raise ValueError('未找到学号列，请确保表头包含: 学号 / student_id / studentid / id')

        data = []
        for row in rows[1:]:
            if not row:
                continue
            raw_id = row[student_id_index]
            student_id = str(raw_id).strip() if raw_id is not None else ''
            if student_id.isdigit():
                data.append((student_id, '123456'))
                print(f'添加学生: {student_id}')

        if data:
            sql = 'INSERT INTO users (username, password) VALUES (%s, %s)'
            cursor.executemany(sql, data)
            conn.commit()
            print(f'✅ 成功插入 {len(data)} 条记录')
        else:
            print('⚠️ 没有找到有效数据')

    except FileNotFoundError:
        print(f'❌ 文件未找到，请检查路径: {EXCEL_PATH}')
    except Exception as exc:
        print(f'❌ 导入失败: {exc}')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()


if __name__ == '__main__':
    main()
