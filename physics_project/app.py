import csv
import io
import json
import logging
import math
import os
import random
import re
import time
import uuid
from decimal import Decimal
from datetime import datetime
from functools import wraps
from math import pi, log

import mysql.connector
import numpy as np
import redis
import sympy as sp
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, Response, abort
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__, template_folder='templates', static_folder='static', static_url_path='/static')
app.secret_key = 'your_secret_key_here'
logger = logging.getLogger(__name__)

# Redis 配置
redis_url = os.getenv('REDIS_URL', 'redis://172.17.66.87:6379/0')
redis_client = redis.Redis.from_url(redis_url, decode_responses=True)
POOL_TARGET = 50
POOL_LOW_WATER = 25
POOL_REFILL_BATCH = 10
PROBLEM_TTL_SECONDS = int(os.getenv('PROBLEM_TTL_SECONDS', 900))

TEMPLATE_CACHE = {}
TEMPLATE_CACHE_TS = 0

PROBLEM_POOL_TARGET_SIZE = int(os.getenv('PROBLEM_POOL_TARGET_SIZE', 20))
PROBLEM_POOL_REFILL_BATCH = int(os.getenv('PROBLEM_POOL_REFILL_BATCH', 10))
PROBLEM_TTL_SECONDS = int(os.getenv('PROBLEM_TTL_SECONDS', 900))

# 数据库配置
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '123456',
    'database': 'physics_new3'
}

# 图片上传配置
UPLOAD_FOLDER = 'static/images'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'svg'}
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 头像配置
AVATAR_FOLDER = 'static/avatars'
AVATAR_ALLOWED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.svg'}
DEFAULT_AVATAR = 'default.svg'
DEFAULT_PASSWORD = '123456'


DEFAULT_EXAM_PAPER_NAME = '默认题库'


def get_or_create_default_exam_paper(cursor):
    """获取或创建默认题库。"""
    cursor.execute("SELECT id FROM exam_papers WHERE name = %s LIMIT 1", (DEFAULT_EXAM_PAPER_NAME,))
    row = cursor.fetchone()
    if row:
        return row['id'] if isinstance(row, dict) else row[0]

    cursor.execute(
        """
        INSERT INTO exam_papers (name, description, is_enabled)
        VALUES (%s, %s, TRUE)
        """,
        (DEFAULT_EXAM_PAPER_NAME, '系统默认题库')
    )
    return cursor.lastrowid


def get_exam_papers(include_disabled=True):
    conn = get_db_connection()
    if not conn:
        return []
    cursor = conn.cursor(dictionary=True)
    try:
        query = "SELECT id, name, description, is_enabled, created_at FROM exam_papers"
        params = []
        if not include_disabled:
            query += " WHERE is_enabled = TRUE"
        query += " ORDER BY created_at DESC, id DESC"
        cursor.execute(query, params)
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()


def get_enabled_exam_papers():
    return get_exam_papers(include_disabled=False)


def get_exam_paper_by_id(paper_id):
    if not paper_id:
        return None
    conn = get_db_connection()
    if not conn:
        return None
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, name, description, is_enabled, created_at FROM exam_papers WHERE id = %s", (paper_id,))
        return cursor.fetchone()
    finally:
        cursor.close()
        conn.close()


def resolve_selected_exam_paper_id(preferred_paper_id=None, include_disabled_for_admin=False):
    """解析当前用户选中的题库。"""
    enabled_papers = get_enabled_exam_papers()
    enabled_ids = {paper['id'] for paper in enabled_papers}
    paper_id = preferred_paper_id or session.get('selected_exam_paper_id')

    if paper_id:
        try:
            paper_id = int(paper_id)
        except (TypeError, ValueError):
            paper_id = None

    if paper_id and paper_id in enabled_ids:
        session['selected_exam_paper_id'] = paper_id
        return paper_id

    if paper_id and include_disabled_for_admin and session.get('username') == 'admin':
        paper = get_exam_paper_by_id(paper_id)
        if paper:
            session['selected_exam_paper_id'] = paper['id']
            return paper['id']

    if enabled_papers:
        session['selected_exam_paper_id'] = enabled_papers[0]['id']
        return enabled_papers[0]['id']

    session.pop('selected_exam_paper_id', None)
    return None


def get_selected_exam_paper(include_disabled_for_admin=False):
    paper_id = resolve_selected_exam_paper_id(include_disabled_for_admin=include_disabled_for_admin)
    if not paper_id:
        return None
    return get_exam_paper_by_id(paper_id)


def persist_selected_exam_paper_id(user_id, paper_id):
    """尽力持久化用户选择的题库；旧库缺列时降级为仅写入会话。"""
    conn = get_db_connection()
    if not conn:
        return False

    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE users SET selected_paper_id = %s WHERE id = %s", (paper_id, user_id))
        conn.commit()
        return True
    except mysql.connector.Error as err:
        if getattr(err, 'errno', None) == 1054:
            logger.warning("users.selected_paper_id 列不存在，跳过持久化题库选择: %s", err)
            conn.rollback()
            return False
        raise
    finally:
        cursor.close()
        conn.close()


def get_problem_templates_by_paper(paper_id=None, enabled_only=True):
    """获取题目模板列表，可按题库和启用状态过滤。"""
    conn = get_db_connection()
    if not conn:
        return []
    cursor = conn.cursor(dictionary=True)
    try:
        query = ["SELECT id, template_name, paper_id FROM problem_templates"]
        conditions = []
        params = []

        if paper_id is not None:
            conditions.append("paper_id = %s")
            params.append(paper_id)
        elif enabled_only:
            enabled_paper_ids = get_enabled_exam_paper_ids()
            if not enabled_paper_ids:
                return []
            placeholders = ', '.join(['%s'] * len(enabled_paper_ids))
            conditions.append(f"paper_id IN ({placeholders})")
            params.extend(enabled_paper_ids)

        if conditions:
            query.append("WHERE " + " AND ".join(conditions))
        query.append("ORDER BY id")
        cursor.execute(" ".join(query), params)
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()


def get_enabled_exam_paper_ids():
    """返回所有已开启题库 ID。"""
    return [paper['id'] for paper in get_enabled_exam_papers()]


def build_enabled_paper_filter(alias, selected_paper_id=None):
    """构建仅统计已开启题库的 SQL 过滤条件。"""
    enabled_paper_ids = get_enabled_exam_paper_ids()

    if selected_paper_id is not None:
        if selected_paper_id in enabled_paper_ids:
            return f" AND {alias}.paper_id = %s", [selected_paper_id]
        return " AND 1 = 0", []

    if not enabled_paper_ids:
        return " AND 1 = 0", []

    placeholders = ', '.join(['%s'] * len(enabled_paper_ids))
    return f" AND {alias}.paper_id IN ({placeholders})", enabled_paper_ids


def get_exam_paper_stats(paper_id):
    conn = get_db_connection()
    if not conn:
        return {'completed_count': 0, 'completed_all': False, 'total_time': 0}
    cursor = conn.cursor(dictionary=True)
    try:
        total_problems = get_total_problem_count(paper_id)
        cursor.execute(
            """
            SELECT COUNT(DISTINCT template_id) AS completed_count, COALESCE(SUM(time_taken), 0) AS total_time
            FROM user_responses
            WHERE user_id = %s AND paper_id = %s AND is_correct = TRUE
              AND (template_id, attempt_count) IN (
                  SELECT template_id, attempt_count
                  FROM user_responses
                  WHERE user_id = %s AND paper_id = %s
                  GROUP BY template_id, attempt_count
                  HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
              )
            """,
            (session['user_id'], paper_id, session['user_id'], paper_id)
        )
        row = cursor.fetchone() or {}
        completed_count = int(row.get('completed_count') or 0)
        return {
            'completed_count': completed_count,
            'completed_all': total_problems > 0 and completed_count >= total_problems,
            'total_time': float(row.get('total_time') or 0),
        }
    finally:
        cursor.close()
        conn.close()


def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_excel_file(filename):
    """检查是否为支持的 Excel 文件"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'xlsx'


def save_uploaded_file(file):
    """保存上传的文件"""
    if file and file.filename != '' and allowed_file(file.filename):
        # 生成安全的文件名
        filename = secure_filename(file.filename)
        # 确保文件名唯一
        base, ext = os.path.splitext(filename)
        counter = 1
        while os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], filename)):
            filename = f"{base}_{counter}{ext}"
            counter += 1

        # 保存文件
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        return filename
    return None


def get_avatar_choices():
    avatars_path = os.path.join(app.root_path, AVATAR_FOLDER)
    if not os.path.isdir(avatars_path):
        return [DEFAULT_AVATAR]
    avatars = []
    for filename in os.listdir(avatars_path):
        _, ext = os.path.splitext(filename)
        if ext.lower() in AVATAR_ALLOWED_EXTENSIONS:
            avatars.append(filename)
    if DEFAULT_AVATAR not in avatars:
        avatars.append(DEFAULT_AVATAR)
    return sorted(set(avatars))


def is_password_hash(value):
    if not value:
        return False
    return value.startswith('pbkdf2:') or value.startswith('scrypt:') or value.startswith('argon2:')


def get_display_name(user):
    name = (user.get('name') if isinstance(user, dict) else None) or ''
    name = str(name).strip()
    return name or user.get('username')


def get_db_connection():
    """获取数据库连接"""
    try:
        conn = mysql.connector.connect(**db_config)
        return conn
    except mysql.connector.Error as err:
        print(f"数据库连接失败：{err}")
        return None


def get_pool_key(template_id):
    return f"exam:pool:{template_id}"


def get_problem_key(token):
    return f"exam:problem:{token}"


def cache_problem_with_token(token, problem_data):
    """将题目数据写入Redis并设置TTL"""
    redis_client.setex(get_problem_key(token), PROBLEM_TTL_SECONDS, json.dumps(problem_data))


def get_problem_by_token(token):
    """通过token从Redis获取题目数据"""
    if not token:
        return None
    raw = redis_client.get(get_problem_key(token))
    if not raw:
        return None
    try:
        redis_client.expire(get_problem_key(token), PROBLEM_TTL_SECONDS)
        return json.loads(raw)
    except json.JSONDecodeError:
        return None


def load_template_from_db(template_id):
    conn = get_db_connection()
    if not conn:
        return None
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM problem_templates WHERE id = %s", (template_id,))
    template = cursor.fetchone()
    cursor.close()
    conn.close()
    return template


def get_template(template_id):
    global TEMPLATE_CACHE, TEMPLATE_CACHE_TS
    if template_id in TEMPLATE_CACHE:
        return TEMPLATE_CACHE[template_id]

    template = load_template_from_db(template_id)
    if template:
        TEMPLATE_CACHE[template_id] = template
        TEMPLATE_CACHE_TS = time.time()
    return template


def refill_problem_pool(template_id, count):
    """生成题目并补充到池中"""
    created = 0
    for _ in range(count):
        problem_data = generate_problem_from_template(template_id)
        if problem_data:
            redis_client.lpush(get_pool_key(template_id), json.dumps(problem_data))
            created += 1
    return created


def ensure_problem_pool(template_id):
    """低水位补货（小批量）"""
    current_size = redis_client.llen(get_pool_key(template_id))
    if current_size < POOL_LOW_WATER:
        refill_problem_pool(template_id, POOL_REFILL_BATCH)


def fetch_problem_from_pool(template_id):
    """从池中获取题目，如果不足则补充"""
    ensure_problem_pool(template_id)
    raw_problem = redis_client.rpop(get_pool_key(template_id))

    if not raw_problem:
        problem_data = generate_problem_from_template(template_id)
        if not problem_data:
            return None, None
        token = uuid.uuid4().hex
        cache_problem_with_token(token, problem_data)
        return token, problem_data

    if not raw_problem:
        return None, None

    try:
        problem_data = json.loads(raw_problem)
    except json.JSONDecodeError:
        return None, None

    token = uuid.uuid4().hex
    cache_problem_with_token(token, problem_data)
    return token, problem_data


def generate_and_cache_problem(template_id):
    """直接生成题目并写入Redis，作为池为空时的兜底"""
    problem_data = generate_problem_from_template(template_id)
    if not problem_data:
        return None, None
    token = uuid.uuid4().hex
    cache_problem_with_token(token, problem_data)
    return token, problem_data


def prewarm_pools():
    print("[PREWARM] 开始预热题目池")
    conn = get_db_connection()
    if not conn:
        print("[PREWARM] 数据库连接失败，跳过预热")
        return

    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id FROM problem_templates")
    template_ids = [row['id'] for row in cursor.fetchall()]
    cursor.close()
    conn.close()

    for template_id in template_ids:
        pool_size = redis_client.llen(get_pool_key(template_id))
        if pool_size < POOL_TARGET:
            to_add = POOL_TARGET - pool_size
            print(f"[PREWARM] 模板 {template_id} 补货 {to_add} 道")
            refill_problem_pool(template_id, to_add)
        else:
            print(f"[PREWARM] 模板 {template_id} 池已满足，当前 {pool_size}")
    print("[PREWARM] 预热完成")


# 登录装饰器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('请先登录！', 'danger')
            return redirect(url_for('login'))
        if session.get('username') != 'admin':
            conn = get_db_connection()
            if not conn:
                flash('数据库连接失败', 'danger')
                return redirect(url_for('login'))
            cursor = conn.cursor(dictionary=True)
            try:
                cursor.execute("SELECT current_session_token FROM users WHERE id = %s", (session['user_id'],))
                user = cursor.fetchone()
                if not user or user['current_session_token'] != session.get('session_token'):
                    session.clear()
                    flash('账号已在其他设备登录', 'danger')
                    return redirect(url_for('login'))
            finally:
                cursor.close()
                conn.close()
        return f(*args, **kwargs)

    return decorated_function


def is_correct(user_answer, correct_answer):
    """判断答案是否正确（允许1%误差，支持科学计数法范围）"""
    logger.debug("is_correct 输入: user_answer=%r, correct_answer=%r", user_answer, correct_answer)

    # 处理None值
    if user_answer is None or correct_answer is None:
        logger.debug("is_correct 答案为空，返回 False")
        return False

    # 转换为浮点数
    try:
        user_float = float(user_answer)
        correct_float = float(correct_answer)
    except (ValueError, TypeError) as e:
        logger.debug("is_correct 数值转换失败: %s", e)
        return False

    # 处理特殊情况：两个都是0
    if user_float == 0 and correct_float == 0:
        logger.debug("is_correct 两者均为0，返回 True")
        return True

    # 处理特殊情况：其中一个为0，另一个不为0
    if user_float == 0 or correct_float == 0:
        # 如果其中一个为0，则要求完全相等（因为0的1%还是0）
        result = user_float == correct_float
        logger.debug("is_correct 存在0值，直接比较结果: %s", result)
        return result

    # 计算相对误差（百分比）
    relative_error = abs((user_float - correct_float) / correct_float) * 100

    # 动态误差阈值（根据数量级调整）
    base_tolerance = 1.0  # 基础1%误差

    # 对于非常大或非常小的数，稍微放宽误差限制
    magnitude = math.log10(abs(correct_float))
    if magnitude > 10:  # 大于10^10
        adjusted_tolerance = min(base_tolerance * 1.5, 2.0)
    elif magnitude < -10:  # 小于10^-10
        adjusted_tolerance = min(base_tolerance * 1.5, 2.0)
    else:
        adjusted_tolerance = base_tolerance

    # 检查相对误差
    result = relative_error <= adjusted_tolerance

    logger.debug(
        "is_correct 计算结果: user=%s correct=%s relative_error=%.6f%% tolerance=%s result=%s",
        f"{user_float:.2e}",
        f"{correct_float:.2e}",
        relative_error,
        adjusted_tolerance,
        result
    )

    return result


# 新增辅助函数：科学计数法格式化
def format_scientific(value, precision=2):
    """将数值格式化为科学计数法字符串"""
    if value == 0:
        return "0.00 × 10⁰"

    is_negative = value < 0
    abs_value = abs(value)

    # 计算指数
    exponent = math.floor(math.log10(abs_value))
    mantissa = abs_value / (10 ** exponent)

    # 调整到标准形式 (1 ≤ mantissa < 10)
    if mantissa >= 10:
        mantissa /= 10
        exponent += 1
    elif mantissa < 1:
        mantissa *= 10
        exponent -= 1

    # 格式化
    sign = '-' if is_negative else ''
    mantissa_str = f"{mantissa:.{precision}f}"

    # 获取上标数字
    superscript_digits = {
        '0': '⁰', '1': '¹', '2': '²', '3': '³', '4': '⁴',
        '5': '⁵', '6': '⁶', '7': '⁷', '8': '⁸', '9': '⁹'
    }

    exp_str = str(exponent)
    sup_exp = ''
    if exp_str[0] == '-':
        sup_exp += '⁻'
        exp_str = exp_str[1:]
    for digit in exp_str:
        sup_exp += superscript_digits.get(digit, digit)

    return f"{sign}{mantissa_str} × 10{sup_exp}"


# 新增函数：用于生成正确答案的科学计数法提示
def get_scientific_hint(correct_answers):
    """生成科学计数法格式的正确答案提示"""
    if not correct_answers or not isinstance(correct_answers, list):
        return "正确答案: 暂无"

    formatted_answers = []
    for answer in correct_answers:
        try:
            formatted = format_scientific(float(answer))
            formatted_answers.append(formatted)
        except:
            formatted_answers.append(str(answer))

    if len(formatted_answers) == 1:
        return f"正确答案: {formatted_answers[0]}"
    else:
        parts = []
        for i, answer in enumerate(formatted_answers):
            parts.append(f"答案{i + 1} = {answer}")
        return f"正确答案: {', '.join(parts)}"


def get_problem_display_info(paper_id=None, enabled_only=True):
    """获取题目的显示信息（支持按题库隔离显示序号）。"""
    templates = get_problem_templates_by_paper(paper_id=paper_id, enabled_only=enabled_only)

    display_mapping = {}
    for display_number, template in enumerate(templates, 1):
        display_mapping[template['id']] = {
            'display_number': display_number,
            'template_name': template['template_name'],
            'actual_id': template['id'],
            'paper_id': template.get('paper_id')
        }

    return display_mapping


def build_display_to_actual_map(paper_id=None):
    """生成显示序号到实际ID的映射，便于前端查找"""
    mapping = get_problem_display_info(paper_id)
    display_to_actual = {}

    for actual_id, info in mapping.items():
        display_number = info['display_number']
        display_to_actual[display_number] = actual_id

    return display_to_actual


def get_display_number(actual_id, paper_id=None):
    """根据实际ID获取显示序号"""
    mapping = get_problem_display_info(paper_id)
    if actual_id in mapping:
        return mapping[actual_id]['display_number']
    return actual_id  # 回退到实际ID


def get_actual_id(display_number, paper_id=None):
    """根据显示序号获取实际ID"""
    mapping = get_problem_display_info(paper_id)
    for actual_id, info in mapping.items():
        if info['display_number'] == display_number:
            return actual_id
    return None  # 找不到对应的实际ID


def has_full_correct_attempt(cursor, user_id, template_id):
    """判断是否存在所有答案均正确的作答记录"""
    cursor.execute("""
        SELECT attempt_count
        FROM user_responses
        WHERE user_id = %s AND template_id = %s
        GROUP BY attempt_count
        HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
        LIMIT 1
    """, (user_id, template_id))
    return cursor.fetchone() is not None


def get_latest_full_correct_attempt(cursor, user_id, template_id):
    """获取最近一次所有答案均正确的作答统计"""
    cursor.execute("""
        SELECT attempt_count
        FROM user_responses
        WHERE user_id = %s AND template_id = %s
        GROUP BY attempt_count
        HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
        ORDER BY attempt_count DESC
        LIMIT 1
    """, (user_id, template_id))
    attempt_row = cursor.fetchone()

    if not attempt_row:
        return None

    attempt_count = attempt_row['attempt_count']
    cursor.execute("""
        SELECT time_taken, attempt_count,
               CASE WHEN is_correct THEN 100 ELSE 0 END as score
        FROM user_responses
        WHERE user_id = %s AND template_id = %s AND attempt_count = %s
        LIMIT 1
    """, (user_id, template_id, attempt_count))
    return cursor.fetchone()


def is_problem_completed(user_id, template_id, paper_id=None):
    """检查指定题目是否已被用户正确完成"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        if paper_id is None:
            return has_full_correct_attempt(cursor, user_id, template_id)
        cursor.execute("""
            SELECT attempt_count
            FROM user_responses
            WHERE user_id = %s AND template_id = %s AND paper_id = %s
            GROUP BY attempt_count
            HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
            LIMIT 1
        """, (user_id, template_id, paper_id))
        return cursor.fetchone() is not None
    finally:
        cursor.close()
        conn.close()


def get_latest_attempt_count(user_id, template_id, paper_id=None):
    """获取用户在某题上的最大 attempt_count，避免会话重置导致编号冲突"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        query = """
            SELECT COALESCE(MAX(attempt_count), 0) AS latest_attempt_count
            FROM user_responses
            WHERE user_id = %s AND template_id = %s
        """
        params = [user_id, template_id]
        if paper_id is not None:
            query += " AND paper_id = %s"
            params.append(paper_id)
        cursor.execute(query, params)
        row = cursor.fetchone() or {}
        return int(row.get('latest_attempt_count') or 0)
    finally:
        cursor.close()
        conn.close()


def get_total_problem_count(paper_id=None):
    """动态获取题目总数（仅统计已开启题库）。"""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        enabled_paper_ids = get_enabled_exam_paper_ids()
        if paper_id is not None:
            if paper_id not in enabled_paper_ids:
                return 0
            cursor.execute("SELECT COUNT(*) FROM problem_templates WHERE paper_id = %s", (paper_id,))
        else:
            if not enabled_paper_ids:
                return 0
            placeholders = ', '.join(['%s'] * len(enabled_paper_ids))
            cursor.execute(
                f"SELECT COUNT(*) FROM problem_templates WHERE paper_id IN ({placeholders})",
                enabled_paper_ids
            )
        row = cursor.fetchone()
        return int(row[0] if row else 0)
    finally:
        cursor.close()
        conn.close()


def generate_problem_from_template(template_id, max_attempts=10):
    """从模板生成具体问题 - 完全动态的合理性验证（无学习表），包含答案单位处理"""
    template = get_template(template_id)

    if not template:
        return None

    variables, configured_ranges = parse_variable_specs(template.get('variables', ''))

    # 符号定义
    x, t, h = sp.symbols('x t h')
    local_vars = {
        'x': x, 't': t, 'h': h, 'sp': sp, 'sqrt': sp.sqrt, 'exp': sp.exp,
        'integrate': sp.integrate, 'pi': pi, 'log': log, 'sin': sp.sin, 'cos': sp.cos
    }

    # 内存中的自适应范围（不持久化）
    reasonable_ranges = {}
    for var in variables:
        reasonable_ranges[var] = configured_ranges.get(var, get_adaptive_default_range(var))

    answer_units = parse_answer_units(template)
    answer_constraints = infer_answer_constraints(answer_units)

    for attempt in range(max_attempts):
        var_values = {}

        # 使用自适应范围生成变量
        for var in variables:
            min_val, max_val = reasonable_ranges[var]
            # 随着尝试次数增加，逐渐扩大搜索范围
            range_expansion = 1.0 + (attempt * 0.1)  # 每次尝试扩大10%
            expanded_min = max(0.01, min_val / range_expansion)
            expanded_max = max_val * range_expansion

            var_values[var] = round(random.uniform(expanded_min, expanded_max), 2)

        # 格式化问题文本
        problem_content = format_problem_text(template['problem_text'], var_values)

        # 计算答案
        try:
            current_vars = local_vars.copy()
            current_vars.update(var_values)

            correct_answer = eval(template['solution_formula'], {}, current_vars)

            if isinstance(correct_answer, tuple):
                correct_answers = [float(a.evalf()) if hasattr(a, 'evalf') else float(a) for a in correct_answer]
            else:
                correct_answers = [
                    float(correct_answer.evalf()) if hasattr(correct_answer, 'evalf') else float(correct_answer)]

            answer_count = template.get('answer_count', 1)
            if len(correct_answers) != answer_count:
                correct_answers = [correct_answers[0]] * answer_count

            # 动态合理性验证（标准随尝试次数放宽）
            if is_answer_reasonable_dynamic(correct_answers, var_values, attempt, answer_constraints):
                # 在内存中更新合理范围（仅本次运行有效）
                for var, value in var_values.items():
                    current_min, current_max = reasonable_ranges[var]
                    reasonable_ranges[var] = (
                        min(current_min, value * 0.8),  # 稍微扩大下限
                        max(current_max, value * 1.2)  # 稍微扩大上限
                    )

                # 确保答案单位数量与答案数量匹配
                answer_count = template.get('answer_count', 1)
                if len(answer_units) < answer_count:
                    # 如果单位数量不足，用空字符串补齐
                    answer_units.extend([''] * (answer_count - len(answer_units)))
                elif len(answer_units) > answer_count:
                    # 如果单位数量过多，只取前answer_count个
                    answer_units = answer_units[:answer_count]

                # 格式化显示答案（保留适当小数位数）
                formatted_correct_answers = []
                for i, answer in enumerate(correct_answers):
                    # 根据答案大小选择合适的小数位数
                    abs_answer = abs(answer)
                    if abs_answer == 0:
                        formatted_correct_answers.append(0.0)
                    elif abs_answer >= 1000:
                        formatted_correct_answers.append(round(answer, 0))
                    elif abs_answer >= 1:
                        formatted_correct_answers.append(round(answer, 2))
                    elif abs_answer >= 0.01:
                        formatted_correct_answers.append(round(answer, 4))
                    else:
                        formatted_correct_answers.append(round(answer, 6))

                # 构建最终返回数据，包含答案单位
                result_data = {
                    'problem_text': problem_content,
                    'var_values': var_values,
                    'correct_answers': formatted_correct_answers,  # 使用格式化后的答案
                    'answer_units': answer_units,  # 添加答案单位
                    'template_id': template_id,
                    'answer_count': template.get('answer_count', 1),
                    'template_name': template['template_name'],
                    'image_filename': template.get('image_filename')
                }

                # 调试信息
                print(f"✅ 题目生成成功 - 模板: {template['template_name']}")
                print(f"   变量值: {var_values}")
                print(f"   正确答案: {formatted_correct_answers}")
                print(f"   答案单位: {answer_units}")
                print(f"   答案数量: {answer_count}")

                return result_data

        except Exception as e:
            # 计算失败，继续尝试
            print(f"⚠️ 题目生成尝试 {attempt + 1} 失败: {str(e)}")
            continue

    # 最终回退：使用保守但保证成功的方法
    return generate_fallback_problem(template, variables, local_vars, reasonable_ranges, answer_units, answer_constraints)


def generate_fallback_problem(template, variables, local_vars, reasonable_ranges=None, answer_units=None, answer_constraints=None):
    """最终回退方案：使用保守范围生成题目"""
    answer_units = answer_units or parse_answer_units(template)
    answer_constraints = answer_constraints or infer_answer_constraints(answer_units)

    correct_answers = [0.0] * template.get('answer_count', 1)
    problem_content = template['problem_text']

    for attempt in range(5):
        var_values = {}
        for var in variables:
            min_val, max_val = (reasonable_ranges or {}).get(var, (1.0, 3.0))
            var_values[var] = round(random.uniform(min_val, max_val), 2)

        problem_content = format_problem_text(template['problem_text'], var_values)

        try:
            current_vars = local_vars.copy()
            current_vars.update(var_values)
            correct_answer = eval(template['solution_formula'], {}, current_vars)

            if isinstance(correct_answer, tuple):
                current_answers = [float(a.evalf()) if hasattr(a, 'evalf') else float(a) for a in correct_answer]
            else:
                current_answers = [
                    float(correct_answer.evalf()) if hasattr(correct_answer, 'evalf') else float(correct_answer)]

            answer_count = template.get('answer_count', 1)
            if len(current_answers) != answer_count:
                current_answers = [current_answers[0]] * answer_count

            if is_answer_reasonable_dynamic(current_answers, var_values, attempt, answer_constraints):
                correct_answers = current_answers
                break
        except Exception:
            continue
    else:
        var_values = {var: 1.0 for var in variables}
        problem_content = format_problem_text(template['problem_text'], var_values)

    # 确保答案单位数量与答案数量匹配
    answer_count = template.get('answer_count', 1)
    if len(answer_units) < answer_count:
        answer_units.extend([''] * (answer_count - len(answer_units)))
    elif len(answer_units) > answer_count:
        answer_units = answer_units[:answer_count]

    # 格式化答案
    formatted_correct_answers = []
    for answer in correct_answers:
        abs_answer = abs(answer)
        if abs_answer >= 1000:
            formatted_correct_answers.append(round(answer, 0))
        elif abs_answer >= 1:
            formatted_correct_answers.append(round(answer, 2))
        else:
            formatted_correct_answers.append(round(answer, 4))

    result_data = {
        'problem_text': problem_content,
        'var_values': var_values,
        'correct_answers': formatted_correct_answers,
        'answer_units': answer_units,  # 包含答案单位
        'template_id': template['id'],
        'answer_count': template.get('answer_count', 1),
        'template_name': template['template_name'],
        'image_filename': template.get('image_filename')
    }

    print(f"⚠️ 使用回退方案生成题目 - 模板: {template['template_name']}")
    print(f"   变量值: {var_values}")
    print(f"   正确答案: {formatted_correct_answers}")
    print(f"   答案单位: {answer_units}")

    return result_data

def get_adaptive_default_range(var_name):
    """根据变量名特征自适应设置默认范围"""
    var_lower = var_name.lower()

    # 基于命名模式的智能猜测
    if any(char in var_lower for char in ['r', 'a', 'l', 'd', 'x', 'h']):  # 几何尺寸类
        return (0.1, 10.0)
    elif any(char in var_lower for char in ['v', 'u', 'w', 'speed', 'velocity']):  # 速度类
        return (1.0, 50.0)
    elif any(char in var_lower for char in ['b', 'e', 'f', 'field']):  # 场强类
        return (0.1, 5.0)
    elif any(char in var_lower for char in ['i', 'current']):  # 电流类
        return (0.1, 10.0)
    elif any(char in var_lower for char in ['r', 'resistance']):  # 电阻类
        return (1.0, 100.0)
    elif any(char in var_lower for char in ['m', 'mass']):  # 质量类
        return (0.01, 5.0)
    elif any(char in var_lower for char in ['omega', 'ω', 'angular']):  # 角速度
        return (1.0, 20.0)
    elif any(char in var_lower for char in ['dbdt', 'alpha', 'rate']):  # 变化率
        return (0.1, 10.0)
    elif any(char in var_lower for char in ['density']):  # 密度
        return (1000.0, 10000.0)
    else:
        # 通用范围
        return (0.5, 20.0)


def is_answer_reasonable_dynamic(correct_answers, var_values, attempt_num, constraints=None):
    """完全动态的合理性验证"""
    if not correct_answers:
        return False

    constraints = constraints or {}
    min_answer = constraints.get('min_answer')
    max_answer = constraints.get('max_answer')
    non_negative = constraints.get('non_negative', False)

    for answer in correct_answers:
        # 基础检查：必须是有限实数
        if not isinstance(answer, (int, float)) or not np.isfinite(answer):
            return False

        if non_negative and answer < 0:
            return False

        abs_answer = abs(answer)

        # 动态阈值：随着尝试次数增加，逐渐放宽标准
        max_threshold = 1e6 * (1 + attempt_num * 0.2)  # 逐渐放宽上限
        min_threshold = 1e-8 / (1 + attempt_num * 0.2)  # 逐渐放宽下限

        if max_answer is not None:
            max_threshold = min(max_threshold, max_answer)
        if min_answer is not None:
            min_threshold = max(min_threshold, min_answer)

        if abs_answer > max_threshold or (0 < abs_answer < min_threshold):
            return False

        # 检查与输入变量的协调性（标准也动态放宽）
        if not check_dynamic_consistency(answer, var_values, attempt_num):
            return False

    return True


def check_dynamic_consistency(answer, var_values, attempt_num):
    """动态检查答案与变量的协调性"""
    if not var_values:
        return True

    abs_answer = abs(answer)
    var_values_list = [abs(v) for v in var_values.values() if isinstance(v, (int, float))]

    if not var_values_list:
        return True

    avg_var = sum(var_values_list) / len(var_values_list)

    # 动态比率阈值：随着尝试次数放宽
    base_max_ratio = 1000
    base_min_ratio = 0.001
    relaxation_factor = 1 + (attempt_num * 0.3)  # 每次尝试放宽30%

    max_ratio = base_max_ratio * relaxation_factor
    min_ratio = base_min_ratio / relaxation_factor

    ratio_to_avg = abs_answer / avg_var if avg_var > 0 else abs_answer

    # 检查是否在合理比率范围内
    if ratio_to_avg > max_ratio or ratio_to_avg < min_ratio:
        return False

    return True


def format_problem_text(problem_text, var_values):
    """格式化问题文本"""
    pattern = r'__(\w+)__'

    def replace_var(match):
        var_name = match.group(1)
        return str(var_values.get(var_name, match.group(0)))

    return re.sub(pattern, replace_var, problem_text)



def parse_variable_specs(variables_text):
    """支持 `v[1,5],a[0.1,2],t` 的轻量范围语法；未配置时沿用默认范围。"""
    if not variables_text:
        return [], {}

    tokens = []
    current = []
    bracket_depth = 0
    for ch in variables_text:
        if ch == '[':
            bracket_depth += 1
        elif ch == ']':
            bracket_depth = max(0, bracket_depth - 1)

        if ch == ',' and bracket_depth == 0:
            token = ''.join(current).strip()
            if token:
                tokens.append(token)
            current = []
            continue

        current.append(ch)

    tail = ''.join(current).strip()
    if tail:
        tokens.append(tail)

    variables = []
    ranges = {}
    for token in tokens:
        match = re.match(r'^([A-Za-z_][A-Za-z0-9_]*)\[\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*\]$', token)
        if match:
            name = match.group(1)
            min_val = float(match.group(2))
            max_val = float(match.group(3))
            if min_val > max_val:
                min_val, max_val = max_val, min_val
            variables.append(name)
            ranges[name] = (min_val, max_val)
        else:
            variables.append(token)

    return variables, ranges


def parse_answer_units(template):
    if template.get('answer_units'):
        return [u.strip() for u in template['answer_units'].split(',')]
    return []


def infer_answer_constraints(answer_units):
    """基于单位做基础安全约束，避免明显不合理答案。"""
    merged_units = ' '.join(answer_units).lower()
    non_negative_units = ['kg', 'j', 'n', 'pa', 'w', 'hz', 'ω', 'ohm', 'Ω']
    non_negative = any(unit.lower() in merged_units for unit in non_negative_units)

    return {
        'non_negative': non_negative,
        'min_answer': 0 if non_negative else None,
        'max_answer': 1e7
    }

def classify_error_type(user_answer, correct_answer, is_correct):
    """根据用户答案与正确答案的差异推断错误类型"""
    if is_correct:
        return '正确'

    if user_answer is None:
        return '未作答'

    try:
        user_float = float(user_answer)
        correct_float = float(correct_answer)
    except (TypeError, ValueError):
        return '格式错误'

    if correct_float == 0:
        return '计算错误' if user_float != 0 else '正确'

    relative_error = abs((user_float - correct_float) / correct_float) * 100

    if relative_error > 50:
        return '概念错误'
    if relative_error > 5:
        return '计算误差'
    return '精度或单位偏差'


def save_user_response(user_id, template_id, paper_id, problem_text, user_answers, correct_answers, is_correct_list,
                       attempt_count, time_taken, error_types=None):
    """保存用户答题记录（支持多答案）"""
    print(f"\n=== 保存答题记录开始 ===")
    print(f"用户ID: {user_id}")
    print(f"模板ID: {template_id}")
    print(f"用户答案: {user_answers}")
    print(f"正确答案: {correct_answers}")
    print(f"是否正确: {is_correct_list}")
    print(f"尝试次数: {attempt_count}")
    print(f"用时: {time_taken}秒")
    print(f"答案数量: {len(user_answers)}")

    conn = None
    try:
        # 1. 获取数据库连接
        conn = get_db_connection()
        if not conn:
            print("❌ 数据库连接失败")
            return False

        cursor = conn.cursor()
        print("✅ 数据库连接成功")

        # 2. 验证用户和模板是否存在
        cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
        user_exists = cursor.fetchone()
        if not user_exists:
            print(f"❌ 用户ID {user_id} 不存在")
            return False
        print("✅ 用户存在")

        cursor.execute("SELECT id FROM problem_templates WHERE id = %s", (template_id,))
        template_exists = cursor.fetchone()
        if not template_exists:
            print(f"❌ 模板ID {template_id} 不存在")
            return False
        print("✅ 模板存在")

        # 3. 截断过长的problem_text
        truncated_problem_text = problem_text[:1000] + "..." if len(problem_text) > 1000 else problem_text
        print(f"✅ 问题文本已截断: {len(truncated_problem_text)} 字符")

        # 4. 保存每个答案的记录
        all_success = True
        saved_count = 0

        if error_types is None:
            error_types = [
                classify_error_type(user_answer, correct_answer, is_correct)
                for user_answer, correct_answer, is_correct in zip(user_answers, correct_answers, is_correct_list)
            ]

        for i, (user_answer, correct_answer, is_correct, error_type) in enumerate(
                zip(user_answers, correct_answers, is_correct_list, error_types)):
            try:
                print(
                    f"正在保存答案 {i + 1}: user_answer={user_answer}, correct_answer={correct_answer}, is_correct={is_correct}, error_type={error_type}")

                cursor.execute("""
                    INSERT INTO user_responses
                    (user_id, template_id, problem_text, user_answer,
                     correct_answer, is_correct, error_type, attempt_count, time_taken, answer_index)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (user_id, template_id, truncated_problem_text, user_answer,
                      correct_answer, is_correct, error_type, attempt_count, time_taken, i))

                saved_count += 1
                print(f"✅ 答案 {i + 1} 保存成功")

            except mysql.connector.Error as err:
                print(f"❌ 答案 {i + 1} 保存失败: {err}")
                all_success = False
                break
            except Exception as e:
                print(f"❌ 答案 {i + 1} 保存异常: {str(e)}")
                all_success = False
                break

        # 5. 提交事务
        if all_success:
            conn.commit()
            print(f"✅ 事务提交成功，共保存 {saved_count} 个答案记录")
        else:
            conn.rollback()
            print("❌ 部分答案保存失败，事务已回滚")

        return all_success

    except mysql.connector.Error as err:
        print(f"❌ 数据库错误: {err}")
        print(f"错误代码: {err.errno}")
        print(f"SQL状态: {err.sqlstate}")
        if conn:
            conn.rollback()
        return False

    except Exception as e:
        print(f"❌ 保存过程中发生异常: {str(e)}")
        import traceback
        print(f"详细错误信息:\n{traceback.format_exc()}")
        if conn:
            conn.rollback()
        return False

    finally:
        if conn and conn.is_connected():
            conn.close()
            print("✅ 数据库连接已关闭")
        print("=== 保存答题记录结束 ===\n")


def repair_database():
    """修复数据库表结构"""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # 检查并添加缺失的列
        cursor.execute("SHOW COLUMNS FROM user_responses LIKE 'user_id'")
        if not cursor.fetchone():
            cursor.execute("ALTER TABLE user_responses ADD COLUMN user_id INT NOT NULL AFTER id")
            print("已添加 user_id 列")

        cursor.execute("SHOW COLUMNS FROM user_responses LIKE 'template_id'")
        if not cursor.fetchone():
            cursor.execute("ALTER TABLE user_responses ADD COLUMN template_id INT NOT NULL AFTER user_id")
            print("已添加 template_id 列")

        cursor.execute("SHOW TABLES LIKE 'exam_papers'")
        if not cursor.fetchone():
            cursor.execute("""
                CREATE TABLE exam_papers (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(100) NOT NULL UNIQUE,
                    description TEXT DEFAULT NULL,
                    is_enabled BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            print("已创建 exam_papers 表")

        cursor = conn.cursor(dictionary=True)
        default_paper_id = get_or_create_default_exam_paper(cursor)
        cursor.close()
        cursor = conn.cursor()

        cursor.execute("SHOW COLUMNS FROM problem_templates LIKE 'paper_id'")
        if not cursor.fetchone():
            cursor.execute("ALTER TABLE problem_templates ADD COLUMN paper_id INT DEFAULT NULL")
            print("已添加 problem_templates.paper_id 列")
        cursor.execute("UPDATE problem_templates SET paper_id = %s WHERE paper_id IS NULL", (default_paper_id,))

        cursor.execute("SHOW COLUMNS FROM user_responses LIKE 'paper_id'")
        if not cursor.fetchone():
            cursor.execute("ALTER TABLE user_responses ADD COLUMN paper_id INT DEFAULT NULL AFTER answer_index")
            print("已添加 user_responses.paper_id 列")
        cursor.execute("""
            UPDATE user_responses ur
            JOIN problem_templates pt ON pt.id = ur.template_id
            SET ur.paper_id = pt.paper_id
            WHERE ur.paper_id IS NULL
        """)

        cursor.execute("SHOW COLUMNS FROM user_responses LIKE 'error_type'")
        if not cursor.fetchone():
            cursor.execute("ALTER TABLE user_responses ADD COLUMN error_type VARCHAR(50) DEFAULT '未知' AFTER is_correct")
            cursor.execute("UPDATE user_responses SET error_type = '正确' WHERE is_correct = TRUE")
            cursor.execute("UPDATE user_responses SET error_type = '计算误差' WHERE is_correct = FALSE")
            print("已添加 error_type 列")

        # 添加外键约束（如果不存在）
        cursor.execute("""
            SELECT COUNT(*) FROM information_schema.table_constraints
            WHERE table_name = 'user_responses' 
            AND constraint_name = 'user_responses_ibfk_1'
        """)
        if cursor.fetchone()[0] == 0:
            cursor.execute("""
                ALTER TABLE user_responses
                ADD FOREIGN KEY (user_id) REFERENCES users(id)
            """)
            print("已添加 user_id 外键约束")

        cursor.execute("""
            SELECT COUNT(*) FROM information_schema.table_constraints
            WHERE table_name = 'user_responses' 
            AND constraint_name = 'user_responses_ibfk_2'
        """)
        if cursor.fetchone()[0] == 0:
            cursor.execute("""
                ALTER TABLE user_responses
                ADD FOREIGN KEY (template_id) REFERENCES problem_templates(id)
            """)
            print("已添加 template_id 外键约束")

        conn.commit()
    except mysql.connector.Error as err:
        print(f"修复数据库失败: {err}")
        conn.rollback()
    finally:
        cursor.close()
        conn.close()


def initialize_database():
    """初始化数据库 - 使用新的图片管理方式，包含答案单位字段"""
    conn = get_db_connection()
    cursor = conn.cursor()

    # 创建用户表（如果不存在）- 添加完成状态字段
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INT AUTO_INCREMENT PRIMARY KEY,
        username VARCHAR(50) NOT NULL UNIQUE,
        name VARCHAR(100) DEFAULT NULL,
        major VARCHAR(100) DEFAULT NULL,
        class_name VARCHAR(100) DEFAULT NULL,
        password VARCHAR(255) NOT NULL,
        avatar_filename VARCHAR(255) DEFAULT 'default.svg',
        password_changed BOOLEAN DEFAULT TRUE,
        current_session_token VARCHAR(64) DEFAULT NULL,
        completed_all BOOLEAN DEFAULT FALSE,
        completed_at TIMESTAMP NULL,
        total_score INT DEFAULT 0,
        total_time FLOAT DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS exam_papers (
        id INT AUTO_INCREMENT PRIMARY KEY,
        name VARCHAR(100) NOT NULL UNIQUE,
        description TEXT DEFAULT NULL,
        is_enabled BOOLEAN DEFAULT TRUE,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    default_paper_id = get_or_create_default_exam_paper(cursor)

    # 创建问题模板表（如果不存在）- 添加图片文件名字段和答案单位字段
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS problem_templates (
        id INT AUTO_INCREMENT PRIMARY KEY,
        template_name VARCHAR(100) NOT NULL,
        problem_text TEXT NOT NULL,
        variables TEXT NOT NULL,
        solution_formula TEXT NOT NULL,
        answer_count INT DEFAULT 1,
        answer_units TEXT,  -- 新增：答案单位字段
        difficulty VARCHAR(20) DEFAULT 'medium',
        image_filename VARCHAR(255) NULL,
        paper_id INT DEFAULT NULL,
        FOREIGN KEY (paper_id) REFERENCES exam_papers(id)
    )
    """)

    # 创建用户答题记录表（确保包含所有必要字段）
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS user_responses (
        id INT AUTO_INCREMENT PRIMARY KEY,
        user_id INT NOT NULL,
        template_id INT NOT NULL,
        problem_text TEXT NOT NULL,
        user_answer FLOAT NOT NULL,
        correct_answer FLOAT NOT NULL,
        is_correct BOOLEAN NOT NULL,
        error_type VARCHAR(50) DEFAULT '未知',
        attempt_count INT NOT NULL,
        time_taken FLOAT NOT NULL,
        response_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        answer_index INT DEFAULT 0,
        paper_id INT DEFAULT NULL,
        FOREIGN KEY (user_id) REFERENCES users(id),
        FOREIGN KEY (template_id) REFERENCES problem_templates(id),
        FOREIGN KEY (paper_id) REFERENCES exam_papers(id)
    )
    """)

    # 创建防伪记录表
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS verification_records (
        id INT AUTO_INCREMENT PRIMARY KEY,
        user_id INT NOT NULL,
        verification_code VARCHAR(20) NOT NULL UNIQUE,
        verification_data JSON NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )
    """)

    # 插入电磁学题目模板 - 包含答案单位信息
    templates = [
        # 题目1：闭合圆形线圈的感应电流（无图片）
        {
            'name': '闭合圆形线圈的感应电流',
            'text': r"""
                <div class="math-formula">
                    <h5>题目描述：</h5>
                    <p>用导线制成一半径为 \( r = __r__ \, \text{cm} \) 的闭合圆形线圈，其电阻 \( R = __R__ \, \Omega \)，均匀磁场垂直于线圈平面。</p>
                    <p>欲使电路中有一稳定的感应电流 \( i = __i__ \, \text{A} \)，求 \( B \) 的变化率 \( \frac{dB}{dt} \)。</p>
                    <div class="problem-hint-static">
                        <h5>解题提示：</h5>
                        <p>法拉第电磁感应定律：\( \varepsilon = -\frac{d\Phi}{dt} \)</p>
                        <p>磁通量：\( \Phi = B \cdot S = B \cdot \pi r^2 \)</p>
                        <p>感应电流：\( i = \frac{\varepsilon}{R} \)</p>
                    </div>
                </div>
            """,
            'variables': 'r,R,i',
            'formula': "i * R / (pi * (r/100)**2)",
            'answer_count': 1,
            'answer_units': 'T/s',  # 磁感应强度变化率，单位：特斯拉/秒
            'image_filename': None
        },

        # 题目2：高铁电磁感应问题（无图片）
        {
            'name': '高铁电磁感应问题',
            'text': r"""
            <div class="math-formula">
            <h5>题目描述：</h5>
            <p>中国是目前世界上高速铁路运行里程最长的国家，已知"复兴号"高铁长度为 L = __L__ m，车厢高 h = __h__ m，正常行驶速度 v = __v__ km/h。</p>
            <p>假设地面附近地磁场的水平分量约为 B = __B__ μT，将列车视为一整块导体，只考虑地磁场的水平分量。</p>
            <p>则"复兴号"列车在自西向东正常行驶的过程中，求车顶与车底之间的电势差大小。</p>
            <div class="problem-hint-static">
                        <h5>解题提示：</h5>
                <p>1. 速度单位换算：km/h → m/s</p>
                <p>2. 磁场单位换算：μT → T</p>
                <p>3. 导体在磁场中运动产生的感应电动势：ε = BLv</p>
                <p>4. 最终结果单位转换为微伏(μV)：1 V = 10⁶ μV</p>
            </div>
            </div>
            """,
            'variables': 'L,h,v,B',
            'formula': "B * L * (v / 3.6)",
            'answer_count': 1,
            'answer_units': 'μV',  # 电势差，单位：微伏
            'image_filename': None
        },

        # 题目3：等边三角形金属框转动电动势（有图片）
        {
            'name': '等边三角形金属框转动电动势',
            'text': r"""
                <div class="math-formula">
                    <h5>题目描述：</h5>
                    <p>如图所示，等边三角形的金属框，边长为 \( l = __l__ \, \text{m} \)，放在均匀磁场 \( B = __B__ \, \text{T} \) 中。</p>
                    <p>\( ab \) 边平行于磁感强度 \( B \)，当金属框绕 \( ab \) 边以角速度 \( \omega = __omega__ \, \text{rad/s} \) 转动时：</p>
                    <ol>
                        <li>求 \( bc \) 边上沿 \( bc \) 的电动势</li>
                        <li>求 \( ca \) 边上沿 \( ca \) 的电动势</li>
                        <li>求金属框内的总电动势</li>
                    </ol>
                    <div class="problem-hint-static">
                        <h5>解题提示：</h5>
                        <p>动生电动势：\( \varepsilon = \int (\vec{v} \times \vec{B}) \cdot d\vec{l} \)</p>
                        <p>考虑各边的运动情况和磁场方向</p>
                    </div>
                </div>
            """,
            'variables': 'l,B,omega',
            'formula': "(3/8) * B * omega * l**2, -(3/8) * B * omega * l**2,0",
            'answer_count': 3,
            'answer_units': 'V,V,V',  # 三个电动势，单位都是伏特
            'image_filename': 'problem3.png'
        },

        # 题目4：动生电动势与感生电动势（有图片）
        {
            'name': '动生电动势与感生电动势',
            'text': r"""
                <div class="math-formula">
                    <h5>题目描述：</h5>
                    <p>导体 \( AC \) 以速度 \( v = __v__ \, \text{m/s} \) 运动。</p>
                    <p>设 \( AC = __AC__ \, \text{cm} \)，均匀磁场随时间的变化率 \( \frac{dB}{dt} = __dBdt__ \, \text{T/s} \)。</p>
                    <p>某一时刻 \( B = __B__ \, \text{T} \)，\( x = __x__ \, \text{cm} \)，求：</p>
                    <ol>
                        <li>这时动生电动势的大小</li>
                        <li>总感应电动势的大小</li>
                        <li>动生电动势随 \( AC \) 运动的变化趋势（增大填1，减小填-1）</li>
                    </ol>
                    <div class="problem-hint-static">
                        <h5>解题提示：</h5>
                        <p>动生电动势：导体切割磁感线产生</p>
                        <p>感生电动势：磁场变化产生</p>
                        <p>总电动势为两者之和</p>
                    </div>
                </div>
            """,
            'variables': 'v,AC,dBdt,B,x',
            'formula': "B * v * (AC/100), (B * v * (AC/100)) + (dBdt * (x/100) * (AC/100)), 1",
            'answer_count': 3,
            'answer_units': 'V,V,-',  # 前两个是伏特，第三个是无量纲
            'image_filename': 'problem4.png'
        },

        # 题目5：折形金属导线运动电势差（有图片）
        {
            'name': '折形金属导线运动电势差',
            'text': r"""
                <div class="math-formula">
                    <h5>题目描述：</h5>
                    <p>\( aOc \) 为一折成 \( 30^\circ \) 角的金属导线（\( aO = Oc = L = __L__ \, \text{m} \)），位于 \( xy \) 平面中。</p>
                    <p>其中 \( aO \) 段与 \( x \) 轴夹角为 \( 30^\circ \)，\( Oc \) 段与 \( x \) 轴夹角为 \( 30^\circ \)，两段在 \( O \) 点相接。</p>
                    <p>磁感强度为 \( B = __B__ \, \text{T} \) 的匀强磁场垂直于 \( xy \) 平面。</p>
                    <ol>
                        <li>当 \( aOc \) 以速度 \( v = __v__ \, \text{m/s} \) 沿 \( x \) 轴正向运动时，导线上 \( a, c \) 两点间电势差 \( U_{ac} \)</li>
                        <li>当 \( aOc \) 以速度 \( v \) 沿 \( y \) 轴正向运动时，判断 \( a, c \) 两点电势高低（a点高填1，c点高填-1）</li>
                    </ol>
                    <div class="problem-hint-static">
                        <h5>解题提示：</h5>
                        <p>动生电动势公式：\( \varepsilon = \int (\vec{v} \times \vec{B}) \cdot d\vec{l} \)</p>
                        <p>考虑不同运动方向时各段的电动势，注意30度角的影响</p>
                        <p>总电势差为各段电动势的代数和</p>
                    </div>
                </div>
            """,
            'variables': 'L,B,v',
            'formula': "B * v * L /2 , -1",
            'answer_count': 2,
            'answer_units': 'V,-',  # 第一个是伏特，第二个是无量纲
            'image_filename': 'problem5.png'
        },

        # 题目6：磁铁插入线圈的感应现象（无图片）
        {
            'name': '磁铁插入线圈的感应现象',
            'text': r"""
                <div class="math-formula">
                    <h5>题目描述：</h5>
                    <p>将磁铁插入闭合电路线圈，一次是迅速地插入，另一次是缓慢地插入。</p>
                    <ol>
                        <li>两次插入过程中，线圈中感应电荷量是否相同？（相同填1，不同填0）</li>
                        <li>两次插入过程中，手推磁铁所做的功是否相同？（相同填1，不同填0）</li>
                    </ol>
                    <div class="problem-hint-static">
                        <h5>解题提示：</h5>
                        <p>感应电荷量：\( q = \frac{\Delta\Phi}{R} \)</p>
                        <p>做功与功率和时间有关</p>
                    </div>
                </div>
            """,
            'variables': '',
            'formula': "1, 0",
            'answer_count': 2,
            'answer_units': '-,-',  # 两个都是无量纲
            'image_filename': None
        },

        # 题目7：双圆线圈的感应电流（有图片）
        {
            'name': '双圆线圈的感应电流',
            'text': r"""
                <div class="math-formula">
                    <h5>题目描述：</h5>
                    <p>电阻为 \( R = __R__ \, \Omega \) 的闭合线圈折成半径分别为 \( a = __a__ \, \text{cm} \) 和 \( 2a \) 的两个圆，</p>
                    <p>将其置于与两圆平面垂直的匀强磁场内，磁感应强度按 \( B = B_0 \sin(\omega t) \) 的规律变化。</p>
                    <p>已知 \( B_0 = __B0__ \, \text{T} \)，\( \omega = __omega__ \, \text{rad/s} \)，求线圈中感应电流的最大值。</p>
                    <div class="problem-hint-static">
                        <h5>解题提示：</h5>
                        <p>法拉第电磁感应定律</p>
                        <p>总电动势为两个线圈电动势之和</p>
                        <p>感应电流最大值</p>
                    </div>
                </div>
            """,
            'variables': 'R,a,B0,omega',
            'formula': "(pi * omega * B0 / R) * ((a/100)**2 + (2*a/100)**2)",
            'answer_count': 1,
            'answer_units': 'A',  # 电流，单位：安培
            'image_filename': 'problem7.png'
        },
    ]

    # 插入模板到数据库 - 包含答案单位信息
    for template in templates:
        cursor.execute("SELECT id FROM problem_templates WHERE template_name = %s", (template['name'],))
        if not cursor.fetchone():
            # 如果有图片文件名，在problem_text中插入图片HTML
            problem_text = template['text']
            if template.get('image_filename'):
                img_html = f'''
                <div class="text-center mb-3">
                    <img src="/static/images/{template['image_filename']}" 
                         alt="{template['name']}" class="problem-image img-fluid">
                    <div class="image-caption text-muted">图：{template['name']}</div>
                </div>
                '''
                problem_text = img_html + problem_text

            # 插入包含答案单位的数据
            cursor.execute("""
                INSERT INTO problem_templates 
                (template_name, problem_text, variables, solution_formula, 
                 answer_count, answer_units, image_filename, paper_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                template['name'],
                problem_text,
                template['variables'],
                template['formula'],
                template['answer_count'],
                template.get('answer_units', ''),
                template.get('image_filename'),
                default_paper_id
            ))
            print(f"✅ 插入题目: {template['name']}, 答案单位: {template.get('answer_units', '无')}")

    conn.commit()
    cursor.close()
    conn.close()

    print("✅ 数据库初始化完成，所有题目已添加答案单位")


def ensure_user_columns():
    """确保用户表包含必要字段"""
    conn = get_db_connection()
    if not conn:
        print("用户表检查失败：数据库连接失败")
        return
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SHOW COLUMNS FROM users")
        existing_columns = {col['Field'] for col in cursor.fetchall()}
        cursor.execute("SHOW COLUMNS FROM users LIKE 'password'")
        password_column = cursor.fetchone()
        if password_column and 'varchar(255)' not in password_column['Type'].lower():
            cursor.execute("ALTER TABLE users MODIFY COLUMN password VARCHAR(255) NOT NULL")
            print("已扩展 password 列长度")
        if 'name' not in existing_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN name VARCHAR(100) DEFAULT NULL AFTER username")
            print("已添加 name 列")
        if 'major' not in existing_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN major VARCHAR(100) DEFAULT NULL AFTER name")
            print("已添加 major 列")
        if 'class_name' not in existing_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN class_name VARCHAR(100) DEFAULT NULL AFTER major")
            print("已添加 class_name 列")
        if 'avatar_filename' not in existing_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN avatar_filename VARCHAR(255) DEFAULT 'default.svg' AFTER password")
            print("已添加 avatar_filename 列")
        if 'password_changed' not in existing_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN password_changed BOOLEAN DEFAULT TRUE AFTER avatar_filename")
            cursor.execute("UPDATE users SET password_changed = TRUE")
            print("已添加 password_changed 列")
        if 'current_session_token' not in existing_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN current_session_token VARCHAR(64) DEFAULT NULL AFTER password_changed")
            print("已添加 current_session_token 列")
        if 'selected_paper_id' not in existing_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN selected_paper_id INT DEFAULT NULL AFTER current_session_token")
            print("已添加 selected_paper_id 列")
        conn.commit()
    except mysql.connector.Error as err:
        print(f"更新用户表失败: {err}")
        conn.rollback()
    finally:
        cursor.close()
        conn.close()


def create_admin_user():
    """创建管理员用户"""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # 检查管理员用户是否已存在
        cursor.execute("SELECT id FROM users WHERE username = 'admin'")
        if not cursor.fetchone():
            # 创建管理员用户
            admin_password = generate_password_hash('admin123')
            cursor.execute(
                "INSERT INTO users (username, name, password, password_changed, avatar_filename) "
                "VALUES (%s, %s, %s, %s, %s)",
                ('admin', '管理员', admin_password, True, DEFAULT_AVATAR)
            )
            conn.commit()
            print("管理员用户创建成功: admin / admin123")
        else:
            print("管理员用户已存在")
    except Exception as e:
        print(f"创建管理员用户失败: {e}")
    finally:
        cursor.close()
        conn.close()


def ensure_default_images():
    """确保默认图片文件存在"""
    import os

    default_images = {
        'problem3.png': '等边三角形金属框示意图',
        'problem4.png': '导体AC运动示意图',
        'problem5.png': '折形金属导线示意图',
        'problem7.png': '双圆线圈示意图'
    }

    images_path = os.path.join(app.root_path, 'static', 'images')
    os.makedirs(images_path, exist_ok=True)

    # 检查默认图片是否存在
    for filename, description in default_images.items():
        file_path = os.path.join(images_path, filename)
        if not os.path.exists(file_path):
            print(f"⚠️ 注意: 默认图片缺失: {filename} - {description}")
            print(f"请将图片文件放置到: {file_path}")

    print("✅ 默认图片检查完成")


def verify_image_consistency():
    """验证图片一致性"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # 检查所有有图片的题目
    cursor.execute("""
        SELECT id, template_name, image_filename 
        FROM problem_templates 
        WHERE image_filename IS NOT NULL
    """)

    templates_with_images = cursor.fetchall()

    print("=== 图片一致性检查 ===")
    missing_images = []
    for template in templates_with_images:
        image_path = os.path.join(app.config['UPLOAD_FOLDER'], template['image_filename'])
        if os.path.exists(image_path):
            print(f"✅ 题目 '{template['template_name']}' 图片存在: {template['image_filename']}")
        else:
            print(f"❌ 题目 '{template['template_name']}' 图片缺失: {template['image_filename']}")
            missing_images.append({
                'template_name': template['template_name'],
                'image_filename': template['image_filename']
            })

    if missing_images:
        print(f"\n⚠️ 总计缺失 {len(missing_images)} 个图片文件:")
        for missing in missing_images:
            print(f"   - {missing['template_name']}: {missing['image_filename']}")

    cursor.close()
    conn.close()
    return len(missing_images) == 0


def update_user_completion_status(user_id):
    """更新用户完成状态 - 使用动态题目总数"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 动态获取题目总数
        total_problems = get_total_problem_count(selected_paper_id)

        # 检查是否完成所有题目
        cursor.execute("""
            SELECT COUNT(DISTINCT template_id) as completed_count
            FROM (
                SELECT template_id, attempt_count
                FROM user_responses
                WHERE user_id = %s AND paper_id = %s
                GROUP BY template_id, attempt_count
                HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
            ) as completed_attempts
        """, (user_id, selected_paper_id))
        completed_count = cursor.fetchone()['completed_count']

        # 计算总分和总用时（仅统计完全正确的尝试）
        cursor.execute("""
            SELECT
                COUNT(*) as total_score,
                SUM(time_taken) as total_time
            FROM user_responses
            WHERE (user_id, template_id, attempt_count) IN (
                SELECT user_id, template_id, attempt_count
                FROM user_responses
                WHERE user_id = %s AND paper_id = %s
                GROUP BY template_id, attempt_count
                HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
            )
            AND is_correct = TRUE
        """, (user_id, selected_paper_id))
        stats = cursor.fetchone()

        completed_all = completed_count >= total_problems  # 使用动态总数

        # 更新用户表
        if completed_all:
            cursor.execute("""
                UPDATE users 
                SET completed_all = TRUE,
                    completed_at = NOW(),
                    total_score = %s,
                    total_time = %s
                WHERE id = %s
            """, (stats['total_score'] or 0, stats['total_time'] or 0, user_id))
        else:
            cursor.execute("""
                UPDATE users 
                SET completed_all = FALSE,
                    total_score = %s,
                    total_time = %s
                WHERE id = %s
            """, (stats['total_score'] or 0, stats['total_time'] or 0, user_id))

        conn.commit()
        return completed_all

    except Exception as e:
        print(f"更新用户完成状态失败: {e}")
        conn.rollback()
        return False
    finally:
        cursor.close()
        conn.close()


def update_all_users_completion_status():
    """为所有用户刷新完成状态，返回更新数量"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("SELECT id FROM users")
        users = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

    for user in users:
        update_user_completion_status(user['id'])

    return len(users)


def get_completion_stats(paper_id=None):
    """获取完成情况统计（仅统计已开启题库）。"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    total_problems = get_total_problem_count(paper_id)
    response_filter, response_params = build_enabled_paper_filter('ur', paper_id)
    response_filter_ur2, response_params_ur2 = build_enabled_paper_filter('ur2', paper_id)

    cursor.execute(f"""
        WITH completed AS (
            SELECT user_id, COUNT(DISTINCT template_id) AS completed_count, COALESCE(SUM(time_taken), 0) AS total_time
            FROM user_responses ur
            WHERE ur.is_correct = TRUE {response_filter}
              AND (ur.user_id, ur.template_id, ur.attempt_count) IN (
                SELECT user_id, template_id, attempt_count
                FROM user_responses ur2
                WHERE ur2.is_correct IN (TRUE, FALSE) {response_filter_ur2}
                GROUP BY user_id, template_id, attempt_count
                HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
              )
            GROUP BY user_id
        )
        SELECT
            COUNT(*) AS total_students,
            SUM(CASE WHEN COALESCE(c.completed_count, 0) >= %s THEN 1 ELSE 0 END) AS completed_count,
            ROUND(SUM(CASE WHEN COALESCE(c.completed_count, 0) >= %s THEN 1 ELSE 0 END) / COUNT(*) * 100, 1) AS completion_rate,
            AVG(COALESCE(c.completed_count, 0)) AS avg_score,
            AVG(COALESCE(c.total_time, 0)) AS avg_time
        FROM users u
        LEFT JOIN completed c ON c.user_id = u.id
        WHERE u.username != 'admin'
    """, response_params + response_params_ur2 + [total_problems, total_problems])
    stats = cursor.fetchone() or {}

    cursor.execute(f"""
        WITH completed AS (
            SELECT user_id, COUNT(DISTINCT template_id) AS completed_count
            FROM user_responses ur
            WHERE ur.is_correct = TRUE {response_filter}
              AND (ur.user_id, ur.template_id, ur.attempt_count) IN (
                SELECT user_id, template_id, attempt_count
                FROM user_responses ur2
                WHERE ur2.is_correct IN (TRUE, FALSE) {response_filter_ur2}
                GROUP BY user_id, template_id, attempt_count
                HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
              )
            GROUP BY user_id
        )
        SELECT COUNT(*) AS today_completions
        FROM users u
        LEFT JOIN completed c ON c.user_id = u.id
        WHERE u.username != 'admin' AND COALESCE(c.completed_count, 0) >= %s AND DATE(u.completed_at) = CURDATE()
    """, response_params + response_params_ur2 + [total_problems])
    today_stats = cursor.fetchone() or {}
    cursor.close()
    conn.close()
    return {'stats': stats, 'today_stats': today_stats, 'top_students': []}


def get_students_by_completion(completed=True, limit=None, offset=0, paper_id=None):
    """按完成状态获取学生列表（仅统计已开启题库）。"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    total_problems = get_total_problem_count(paper_id)
    response_filter, response_params = build_enabled_paper_filter('ur', paper_id)
    response_filter_ur2, response_params_ur2 = build_enabled_paper_filter('ur2', paper_id)
    query = f"""
        WITH completed AS (
            SELECT user_id, COUNT(DISTINCT template_id) AS total_score, COALESCE(SUM(time_taken), 0) AS total_time
            FROM user_responses ur
            WHERE ur.is_correct = TRUE {response_filter}
              AND (ur.user_id, ur.template_id, ur.attempt_count) IN (
                SELECT user_id, template_id, attempt_count
                FROM user_responses ur2
                WHERE 1=1 {response_filter_ur2}
                GROUP BY user_id, template_id, attempt_count
                HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
              )
            GROUP BY user_id
        )
        SELECT u.id, u.username, u.name, u.major, u.class_name,
               CASE WHEN COALESCE(c.total_score, 0) >= %s AND %s > 0 THEN TRUE ELSE FALSE END AS completed_all,
               u.completed_at, COALESCE(c.total_score, 0) AS total_score, COALESCE(c.total_time, 0) AS total_time, u.created_at
        FROM users u
        LEFT JOIN completed c ON c.user_id = u.id
        WHERE u.username != 'admin'
        HAVING completed_all = %s
        ORDER BY total_score DESC, total_time ASC, created_at ASC
    """
    full_params = response_params + response_params_ur2 + [total_problems, total_problems, completed]
    if limit:
        query += " LIMIT %s OFFSET %s"
        full_params.extend([limit, offset])
    cursor.execute(query, full_params)
    students = cursor.fetchall()
    cursor.close()
    conn.close()
    return students
def get_class_comparison_stats(paper_id=None):
    """获取班级对比统计数据（用于管理端分析）"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    total_problems = get_total_problem_count(paper_id)
    response_filter = ""
    params = []
    if paper_id is not None:
        response_filter = " AND ur.paper_id = %s"
        params.append(paper_id)
    cursor.execute(f"""
        WITH completed AS (
            SELECT user_id, COUNT(DISTINCT template_id) AS total_score, COALESCE(SUM(time_taken), 0) AS total_time
            FROM user_responses ur
            WHERE ur.is_correct = TRUE {response_filter}
              AND (ur.user_id, ur.template_id, ur.attempt_count) IN (
                SELECT user_id, template_id, attempt_count
                FROM user_responses ur2
                WHERE 1=1 {response_filter.replace('ur.', 'ur2.')}
                GROUP BY user_id, template_id, attempt_count
                HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
              )
            GROUP BY user_id
        )
        SELECT COALESCE(NULLIF(TRIM(u.class_name), ''), '未分班') AS class_name,
               COUNT(*) AS student_count,
               SUM(CASE WHEN COALESCE(c.total_score, 0) >= %s AND %s > 0 THEN 1 ELSE 0 END) AS completed_count,
               ROUND(SUM(CASE WHEN COALESCE(c.total_score, 0) >= %s AND %s > 0 THEN 1 ELSE 0 END) / COUNT(*) * 100, 1) AS completion_rate,
               ROUND(AVG(COALESCE(c.total_score, 0)), 1) AS avg_score,
               ROUND(AVG(COALESCE(c.total_time, 0)), 1) AS avg_time
        FROM users u
        LEFT JOIN completed c ON c.user_id = u.id
        WHERE u.username != 'admin'
        GROUP BY COALESCE(NULLIF(TRIM(u.class_name), ''), '未分班')
        ORDER BY completion_rate DESC, avg_score DESC
    """, params + params + [total_problems, total_problems, total_problems, total_problems])
    class_stats = cursor.fetchall()
    cursor.close()
    conn.close()
    return class_stats
def diagnose_database_issue():
    """诊断数据库问题"""
    print("\n=== 数据库诊断开始 ===")

    try:
        conn = get_db_connection()
        if not conn:
            print("❌ 数据库连接失败")
            return False

        cursor = conn.cursor(dictionary=True)

        # 检查表是否存在
        cursor.execute("SHOW TABLES LIKE 'user_responses'")
        if not cursor.fetchone():
            print("❌ user_responses 表不存在")
            return False

        # 检查表结构
        cursor.execute("DESCRIBE user_responses")
        columns = cursor.fetchall()
        print("✅ user_responses 表结构:")
        for col in columns:
            print(f"  - {col['Field']} ({col['Type']})")

        # 检查外键约束
        cursor.execute("""
            SELECT TABLE_NAME, COLUMN_NAME, CONSTRAINT_NAME, REFERENCED_TABLE_NAME, REFERENCED_COLUMN_NAME
            FROM information_schema.KEY_COLUMN_USAGE
            WHERE TABLE_NAME = 'user_responses' AND REFERENCED_TABLE_NAME IS NOT NULL
        """)
        foreign_keys = cursor.fetchall()
        print("✅ 外键约束:")
        for fk in foreign_keys:
            print(f"  - {fk['COLUMN_NAME']} -> {fk['REFERENCED_TABLE_NAME']}.{fk['REFERENCED_COLUMN_NAME']}")

        cursor.close()
        conn.close()
        print("✅ 数据库诊断完成")
        return True

    except Exception as e:
        print(f"❌ 数据库诊断失败: {str(e)}")
        return False


# 用户认证路由
@app.route('/register', methods=['GET', 'POST'])
def register():
    abort(404)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
            user = cursor.fetchone()

            if user and is_password_hash(user['password']):
                password_ok = check_password_hash(user['password'], password)
            else:
                password_ok = user is not None and user['password'] == password

            if user and password_ok:
                if not is_password_hash(user['password']):
                    new_hash = generate_password_hash(password)
                    try:
                        cursor.execute("UPDATE users SET password = %s WHERE id = %s", (new_hash, user['id']))
                        conn.commit()
                        user['password'] = new_hash
                    except mysql.connector.Error as err:
                        conn.rollback()
                        app.logger.warning(
                            "登录时升级密码哈希失败 user_id=%s, error=%s",
                            user['id'],
                            err,
                        )

                session['user_id'] = user['id']
                session['username'] = user['username']
                session['display_name'] = get_display_name(user)
                session['avatar_filename'] = user.get('avatar_filename') or DEFAULT_AVATAR
                session['is_admin'] = (user['username'] == 'admin')

                if user['username'] != 'admin':
                    session_token = uuid.uuid4().hex
                    cursor.execute("UPDATE users SET current_session_token = %s WHERE id = %s", (session_token, user['id']))
                    conn.commit()
                    session['session_token'] = session_token
                else:
                    session.pop('session_token', None)

                if not user.get('password_changed', True):
                    session['show_password_modal'] = True

                flash('登录成功！', 'success')
                return redirect(url_for('dashboard'))

            flash('用户名或密码错误！', 'danger')
        except mysql.connector.Error as err:
            conn.rollback()
            app.logger.error("登录流程数据库错误: %s", err)
            flash('登录失败，请稍后重试。', 'danger')
        finally:
            cursor.close()
            conn.close()

    return render_template('login.html')


@app.route('/logout')
def logout():
    if 'user_id' in session and session.get('username') != 'admin':
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            try:
                cursor.execute("UPDATE users SET current_session_token = NULL WHERE id = %s", (session['user_id'],))
                conn.commit()
            finally:
                cursor.close()
                conn.close()
    session.clear()
    flash('您已成功退出。', 'success')
    return redirect(url_for('login'))


@app.route('/user/password', methods=['POST'])
@login_required
def update_password():
    current_password = request.form.get('current_password', '')
    new_password = request.form.get('new_password', '')
    confirm_password = request.form.get('confirm_password', '')

    if not new_password or new_password != confirm_password:
        flash('新密码与确认密码不一致', 'danger')
        return redirect(request.referrer or url_for('dashboard'))

    conn = get_db_connection()
    if not conn:
        flash('数据库连接失败', 'danger')
        return redirect(request.referrer or url_for('dashboard'))

    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT password FROM users WHERE id = %s", (session['user_id'],))
        user = cursor.fetchone()
        if not user:
            flash('用户不存在', 'danger')
            return redirect(url_for('login'))

        stored_password = user['password']
        if is_password_hash(stored_password):
            password_ok = check_password_hash(stored_password, current_password)
        else:
            password_ok = stored_password == current_password

        if not password_ok:
            flash('当前密码不正确', 'danger')
            return redirect(request.referrer or url_for('dashboard'))

        new_hash = generate_password_hash(new_password)
        cursor.execute(
            "UPDATE users SET password = %s, password_changed = TRUE WHERE id = %s",
            (new_hash, session['user_id'])
        )
        conn.commit()
        session.pop('show_password_modal', None)
        flash('密码修改成功', 'success')
        return redirect(request.referrer or url_for('dashboard'))
    finally:
        cursor.close()
        conn.close()


@app.route('/user/name', methods=['POST'])
@login_required
def update_name():
    new_name = request.form.get('name', '').strip()
    if not new_name:
        flash('姓名不能为空', 'danger')
        return redirect(request.referrer or url_for('dashboard'))

    conn = get_db_connection()
    if not conn:
        flash('数据库连接失败', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE users SET name = %s WHERE id = %s", (new_name, session['user_id']))
        conn.commit()
        session['display_name'] = new_name
        flash('姓名修改成功', 'success')
        return redirect(request.referrer or url_for('dashboard'))
    finally:
        cursor.close()
        conn.close()


@app.route('/user/avatar', methods=['POST'])
@login_required
def update_avatar():
    avatar_filename = request.form.get('avatar_filename', '')
    available_avatars = get_avatar_choices()
    if avatar_filename not in available_avatars:
        flash('请选择有效的头像', 'danger')
        return redirect(request.referrer or url_for('dashboard'))

    conn = get_db_connection()
    if not conn:
        flash('数据库连接失败', 'danger')
        return redirect(request.referrer or url_for('dashboard'))
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE users SET avatar_filename = %s WHERE id = %s", (avatar_filename, session['user_id']))
        conn.commit()
        session['avatar_filename'] = avatar_filename
        flash('头像已更新', 'success')
        return redirect(request.referrer or url_for('dashboard'))
    finally:
        cursor.close()
        conn.close()


# 主应用路由
@app.route('/')
def home():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('home.html')


@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        available_papers = get_enabled_exam_papers()
        selected_paper_id = resolve_selected_exam_paper_id()
        selected_paper = get_exam_paper_by_id(selected_paper_id) if selected_paper_id else None
        has_available_papers = bool(available_papers)

        if not has_available_papers:
            flash('教师没有布置题目。', 'info')
        elif session.get('username') != 'admin' and not selected_paper_id:
            flash('当前暂无可用题库，请联系管理员开启试卷。', 'warning')

        display_mapping = get_problem_display_info(selected_paper_id) if selected_paper_id else {}
        total_problems = len(display_mapping)
        paper_stats = get_exam_paper_stats(selected_paper_id) if selected_paper_id else {'completed_count': 0, 'completed_all': False, 'total_time': 0}
        completed_all = paper_stats['completed_all']
        completed_count = paper_stats['completed_count']

        current_display_number = 1
        if selected_paper_id and total_problems and not completed_all:
            for display_info in display_mapping.values():
                actual_id = display_info['actual_id']
                cursor.execute(
                    """
                    SELECT 1
                    FROM user_responses
                    WHERE user_id = %s AND template_id = %s AND paper_id = %s
                    GROUP BY attempt_count
                    HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
                    LIMIT 1
                    """,
                    (session['user_id'], actual_id, selected_paper_id)
                )
                if not cursor.fetchone():
                    current_display_number = display_info['display_number']
                    break

        session.pop('attempt_count', None)
        session.pop('current_problem', None)

        return render_template(
            'dashboard.html',
            display_name=session.get('display_name', session.get('username')),
            is_admin=session.get('username') == 'admin',
            current_problem=current_display_number,
            completed_count=completed_count,
            completed_all=completed_all,
            total_problems=total_problems,
            display_mapping=display_mapping,
            available_papers=available_papers,
            selected_paper=selected_paper,
            selected_paper_id=selected_paper_id,
            has_available_papers=has_available_papers
        )

    except mysql.connector.Error as err:
        print(f"数据库查询错误: {err}")
        flash('数据库查询错误', 'danger')
        return redirect(url_for('home'))
    finally:
        cursor.close()
        conn.close()


@app.route('/select_exam_paper', methods=['POST'])
@login_required
def select_exam_paper():
    paper_id = request.form.get('paper_id', type=int)
    selected = get_exam_paper_by_id(paper_id)
    if not selected or not selected.get('is_enabled'):
        flash('所选题库不可用，请重新选择。', 'danger')
        return redirect(url_for('dashboard'))

    session['selected_exam_paper_id'] = paper_id
    persist_selected_exam_paper_id(session['user_id'], paper_id)

    flash(f"已切换到题库：{selected['name']}", 'success')
    return redirect(url_for('dashboard'))


@app.route('/stats')
@login_required
def statistics():
    """统计信息页面"""
    conn = None
    try:
        if 'user_id' not in session:
            flash('请先登录', 'warning')
            return redirect(url_for('login'))

        conn = get_db_connection()
        if not conn:
            flash('数据库连接失败', 'danger')
            return redirect(url_for('dashboard'))

        cursor = conn.cursor(dictionary=True)
        selected_paper_id = resolve_selected_exam_paper_id()

        # 总体统计 - 添加错误处理
        cursor.execute("""
            WITH attempt_summary AS (
                SELECT
                    template_id,
                    attempt_count,
                    COUNT(*) AS total_answers,
                    SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) AS correct_answers,
                    MAX(time_taken) AS time_taken
                FROM user_responses
                WHERE user_id = %s AND paper_id = %s
                GROUP BY template_id, attempt_count
            )
            SELECT
                COUNT(*) AS total_count,
                SUM(CASE WHEN correct_answers = total_answers THEN 1 ELSE 0 END) AS correct_count,
                AVG(time_taken) AS avg_time
            FROM attempt_summary
        """, (session['user_id'], selected_paper_id))

        stats_result = cursor.fetchone()

        # 处理空数据的情况
        if not stats_result or stats_result['total_count'] is None:
            stats = {
                'total_count': 0,
                'correct_count': 0,
                'avg_time': 0
            }
        else:
            # 转换Decimal类型为float
            stats = {
                'total_count': int(stats_result['total_count']),
                'correct_count': int(stats_result['correct_count'] or 0),
                'avg_time': float(stats_result['avg_time'] or 0)
            }

        # 各题目统计 - 添加错误处理
        cursor.execute("""
            WITH attempt_summary AS (
                SELECT
                    template_id,
                    attempt_count,
                    COUNT(*) AS total_answers,
                    SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) AS correct_answers,
                    MAX(time_taken) AS time_taken
                FROM user_responses
                WHERE user_id = %s AND paper_id = %s
                GROUP BY template_id, attempt_count
            )
            SELECT
                t.template_name,
                COUNT(*) AS total,
                SUM(CASE WHEN s.correct_answers = s.total_answers THEN 1 ELSE 0 END) AS correct,
                AVG(s.time_taken) AS avg_time
            FROM attempt_summary s
            JOIN problem_templates t ON s.template_id = t.id
            GROUP BY s.template_id, t.template_name
            ORDER BY t.id
        """, (session['user_id'], selected_paper_id))

        problem_stats_result = cursor.fetchall()

        # 转换problem_stats中的Decimal类型
        problem_stats = []
        for stat in problem_stats_result:
            problem_stats.append({
                'template_name': stat['template_name'],
                'total': int(stat['total']),
                'correct': int(stat['correct'] or 0),
                'avg_time': float(stat['avg_time'] or 0)
            })

        # 计算正确率
        accuracy = 0
        if stats['total_count'] > 0:
            accuracy = round(stats['correct_count'] / stats['total_count'] * 100, 1)

        print(
            f"[DEBUG] 统计数据: total_count={stats['total_count']}, correct_count={stats['correct_count']}, accuracy={accuracy}%")
        print(f"[DEBUG] 题目统计: {len(problem_stats)} 条记录")

        return render_template('stats.html',
                               accuracy=accuracy,
                               correct_count=stats['correct_count'],
                               total_count=stats['total_count'],
                               avg_time=round(stats['avg_time'], 1),
                               problem_stats=problem_stats,
                               username=session['username'],
                           selected_paper_id=selected_paper_id)

    except Exception as e:
        print(f"[统计页面错误] {str(e)}")
        import traceback
        print(f"[详细错误] {traceback.format_exc()}")
        flash(f'获取统计信息失败: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))
    finally:
        if conn and conn.is_connected():
            conn.close()


@app.route('/debug/images')
def debug_images():
    import os
    static_path = os.path.join(app.root_path, 'static', 'images')
    files = os.listdir(static_path) if os.path.exists(static_path) else []
    return jsonify({
        'static_folder': app.static_folder,
        'images_path': static_path,
        'files': files
    })


@app.route('/debug/reload_templates', methods=['GET'])
def reload_templates():
    global TEMPLATE_CACHE, TEMPLATE_CACHE_TS
    TEMPLATE_CACHE = {}
    TEMPLATE_CACHE_TS = time.time()
    return jsonify({'success': True, 'message': '模板缓存已清空', 'timestamp': TEMPLATE_CACHE_TS})


@app.route('/history')
@login_required
def history():
    """获取答题历史"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        selected_paper_id = resolve_selected_exam_paper_id()

        # 获取答题记录
        cursor.execute("""
            SELECT 
                r.id,
                t.template_name,
                r.problem_text,
                r.user_answer,
                r.correct_answer,
                r.is_correct,
                r.attempt_count,
                r.response_time,
                r.time_taken,
                t.id as template_id
            FROM user_responses r
            JOIN problem_templates t ON r.template_id = t.id
            WHERE r.user_id = %s AND r.paper_id = %s
            ORDER BY r.response_time DESC
        """, (session['user_id'], selected_paper_id))

        responses = cursor.fetchall()

        # 格式化时间
        for response in responses:
            if response['response_time']:
                response['formatted_time'] = response['response_time'].strftime('%Y-%m-%d %H:%M:%S')
            else:
                response['formatted_time'] = '未知时间'

        # 获取统计信息
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN is_correct = TRUE THEN 1 ELSE 0 END) as correct_count,
                AVG(time_taken) as avg_time
            FROM user_responses 
            WHERE user_id = %s AND paper_id = %s
        """, (session['user_id'], selected_paper_id))

        stats_result = cursor.fetchone()

        # 处理统计信息
        stats = None
        if stats_result and stats_result['total']:
            stats = {
                'total': int(stats_result['total']),
                'correct': int(stats_result['correct_count'] or 0),
                'avg_time': float(stats_result['avg_time'] or 0)
            }

        return render_template('history.html',
                               responses=responses,
                               stats=stats,
                               username=session['username'],
                               selected_paper_id=selected_paper_id)

    except Exception as e:
        print(f"[系统错误] 获取答题历史失败: {str(e)}")
        flash('获取答题历史失败，请稍后再试', 'danger')
        return render_template('history.html',
                               responses=[],
                               stats=None,
                               username=session['username'],
                               selected_paper_id=selected_paper_id)
    finally:
        if conn:
            conn.close()


@app.route('/debug/history')
@login_required
def debug_history():
    """调试历史记录"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # 检查用户记录
    cursor.execute("SELECT COUNT(*) as count FROM user_responses WHERE user_id = %s", (session['user_id'],))
    user_count = cursor.fetchone()

    # 检查表结构
    cursor.execute("DESCRIBE user_responses")
    table_structure = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify({
        'user_id': session['user_id'],
        'user_response_count': user_count['count'],
        'table_structure': table_structure
    })


@app.route('/refresh_problem/<int:problem_id>', methods=['POST'])
@login_required
def refresh_problem(problem_id):
    """刷新单个题目"""
    try:
        # 根据显示序号获取实际ID
        paper_id = resolve_selected_exam_paper_id()
        actual_id = get_actual_id(problem_id, paper_id)
        if actual_id is None:
            return jsonify({'success': False, 'message': '无效的题目编号'})

        token, problem_data = fetch_problem_from_pool(actual_id)
        if not problem_data:
            token, problem_data = generate_and_cache_problem(actual_id)

        if not problem_data or not token:
            return jsonify({'success': False, 'message': '题目生成失败'})

        # 更新session中的题目状态
        if 'current_problem' in session and session['current_problem']['display_number'] == problem_id:
            session['current_problem'].update({
                'token': token,
                'total_attempts': 0,
                'answered_correctly': False,
                'start_time': time.time(),
                'actual_id': actual_id
            })

        return jsonify({'success': True, 'message': '题目刷新成功', 'token': token})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/problem_ajax/<int:problem_id>')  # 保持参数名为 problem_id
@login_required
def problem_ajax(problem_id):
    """支持Ajax的问题页面 - 内部将problem_id作为显示序号使用"""
    print(f"\n=== Ajax问题页面开始 ===")
    print(f"显示序号: {problem_id}")

    # 根据显示序号获取实际ID
    paper_id = resolve_selected_exam_paper_id()
    actual_id = get_actual_id(problem_id, paper_id)
    if actual_id is None:
        flash('无效的题目编号', 'danger')
        return redirect(url_for('dashboard'))

    # 获取题目显示映射和总数
    paper_id = resolve_selected_exam_paper_id()
    display_mapping = get_problem_display_info(paper_id)
    display_to_actual = build_display_to_actual_map(paper_id)
    total_problems = len(display_mapping)

    # 1. 验证题目序号
    if problem_id < 1 or problem_id > total_problems:
        flash('无效的题目编号', 'danger')
        return redirect(url_for('dashboard'))

    # 2. 初始化或获取题目数据
    problem_data = None
    problem_token = None

    if ('current_problem' in session and
            session['current_problem'].get('display_number') == problem_id):
        problem_token = session['current_problem'].get('token')
        problem_data = get_problem_by_token(problem_token)
        if not problem_data:
            print("缓存中的题目已过期，重新获取新题目")

    if not problem_data:
        print(f"生成/获取新题目... 实际ID: {actual_id}")
        problem_token, problem_data = fetch_problem_from_pool(actual_id)

        if not problem_data:
            problem_token, problem_data = generate_and_cache_problem(actual_id)

        if not problem_data:
            flash('题目生成失败', 'danger')
            return redirect(url_for('dashboard'))

        session['current_problem'] = {
            'display_number': problem_id,
            'actual_id': actual_id,
            'token': problem_token,
            'total_attempts': 0,
            'answered_correctly': False,
            'start_time': time.time()
        }
        print(f"新题目生成成功，答案数量: {problem_data.get('answer_count', 1)}")
    else:
        # 确保最小状态存在
        session['current_problem'].setdefault('total_attempts', 0)
        session['current_problem'].setdefault('answered_correctly', False)
        session['current_problem'].setdefault('start_time', time.time())
        session['current_problem'].setdefault('token', problem_token)
        session['current_problem']['actual_id'] = actual_id

    # 3. 检查是否已经完成
    if session['current_problem'].get('answered_correctly', False):
        print(f"题目 {problem_id} 已完成")

    # 4. 渲染Ajax模板
    total_attempts = session['current_problem'].get('total_attempts', 0)
    answer_count = problem_data.get('answer_count', 1)

    # 确保token写回session（兼容旧数据）
    session['current_problem']['token'] = problem_token or session['current_problem'].get('token')
    session.modified = True

    print(f"渲染Ajax模板，显示序号: {problem_id}, 实际ID: {actual_id}")
    print(f"答案数量: {answer_count}, 累计尝试: {total_attempts}")
    print(f"当前题目参数: {problem_data['var_values']}")
    print(f"=== Ajax问题页面结束 ===\n")

    is_completed = is_problem_completed(session['user_id'], actual_id, paper_id)

    return render_template('problem_ajax.html',
                           problem=problem_data,
                           problem_id=problem_id,  # 传递显示序号到模板
                           total_attempts=total_attempts,
                           answer_count=answer_count,
                           username=session['username'],
                           total_problems=total_problems,
                           display_mapping=display_mapping,
                           display_to_actual=display_to_actual,
                           is_completed=is_completed)


@app.route('/api/submit/<int:problem_id>', methods=['POST'])  # 保持参数名为 problem_id
@login_required
def api_submit(problem_id):
    """API接口：提交答案 - 内部将problem_id作为显示序号使用"""
    try:
        data = request.get_json()
        logger.info("提交答案: problem_id=%s", problem_id)
        logger.debug("提交数据详情: problem_id=%s payload=%s", problem_id, data)

        if not data:
            return jsonify({'success': False, 'message': '无效的请求数据'})

        # 根据显示序号获取实际ID
        paper_id = resolve_selected_exam_paper_id()
        actual_id = get_actual_id(problem_id, paper_id)
        if actual_id is None:
            return jsonify({'success': False, 'message': '无效的题目编号'})

        # 获取题目显示映射和总数
        paper_id = resolve_selected_exam_paper_id()
        display_mapping = get_problem_display_info(paper_id)
        total_problems = len(display_mapping)

        # 验证会话
        if ('current_problem' not in session or
                session['current_problem']['display_number'] != problem_id):
            return jsonify({'success': False, 'message': '会话过期，请重新开始答题'})

        problem_token = session['current_problem'].get('token')
        problem_data = get_problem_by_token(problem_token)
        if not problem_data:
            return jsonify({'success': False, 'message': '题目已过期，请刷新后重试'})

        correct_answers = problem_data.get('correct_answers', [])
        answer_count = problem_data.get('answer_count', 1)
        time_taken = float(data.get('time_taken', 0))
        user_id = session['user_id']
        template_id = problem_data['template_id']
        problem_text = problem_data['problem_text']

        # 如果题目已完成，阻止重复作答
        if is_problem_completed(user_id, actual_id, paper_id):
            return jsonify({
                'success': False,
                'message': '该题已完成，无需重复作答',
                'already_completed': True
            })

        logger.info(
            "处理提交: user_id=%s problem_id=%s actual_id=%s template_id=%s answer_count=%s attempts=%s time_taken=%.2fs",
            user_id,
            problem_id,
            actual_id,
            template_id,
            answer_count,
            session['current_problem']['total_attempts'],
            time_taken,
        )
        logger.debug("正确答案详情: %s", correct_answers)

        # 验证答题时间的合理性
        if time_taken < 0:
            logger.warning("无效的答题时间: %.2fs，重置为0", time_taken)
            time_taken = 0
        elif time_taken > 86400:  # 超过24小时
            logger.warning("答题时间异常长: %.2fs，限制为3600秒", time_taken)
            time_taken = 3600
        elif time_taken < 1:  # 少于1秒（可能有问题）
            logger.warning("答题时间过短: %.2fs，可能计时器有问题", time_taken)
        elif time_taken > 3600:  # 超过1小时
            logger.info("答题时间较长: %.2fs", time_taken)

        logger.info("最终记录用时: %.2fs", time_taken)

        # 获取用户答案
        user_answers = []
        for i in range(answer_count):
            if answer_count == 1:
                user_answer = float(data.get('answer', 0))
                user_answers.append(user_answer)
            else:
                user_answer = float(data.get(f'answer{i + 1}', 0))
                user_answers.append(user_answer)

        logger.debug("用户答案详情: %s", user_answers)

        # 验证每个答案
        is_correct_list = []
        all_correct = True

        for i, (user_answer, correct_answer) in enumerate(zip(user_answers, correct_answers)):
            # 使用不同的变量名避免冲突
            answer_is_correct = is_correct(user_answer, correct_answer)
            is_correct_list.append(answer_is_correct)
            if not answer_is_correct:
                all_correct = False
            logger.debug(
                "答案校验: index=%s user=%s correct=%s result=%s",
                i + 1,
                user_answer,
                correct_answer,
                answer_is_correct,
            )

        error_types = [
            classify_error_type(user_answer, correct_answer, is_correct)
            for user_answer, correct_answer, is_correct in zip(user_answers, correct_answers, is_correct_list)
        ]

        # 计算 attempt_count：基于数据库最大值递增，避免会话重置导致 attempt_count 重复
        latest_attempt_count = get_latest_attempt_count(user_id, template_id, paper_id)
        total_attempts = latest_attempt_count + 1
        session['current_problem']['total_attempts'] = total_attempts

        # 保存答题记录 - 使用更新后的累计尝试次数
        save_success = save_user_response(
            user_id, template_id, paper_id, problem_text, user_answers,
            correct_answers, is_correct_list, total_attempts, time_taken, error_types
        )

        if not save_success:
            logger.error("保存答题记录失败: user_id=%s template_id=%s", user_id, template_id)
            # 即使保存失败，也继续处理，但记录警告
            session['current_problem']['save_failed'] = True

        # 生成正确答案消息
        correct_answer_message = "正确答案: "
        if answer_count == 1:
            correct_answer_message += f"{correct_answers[0]:.2f}"
        else:
            correct_parts = []
            for i, correct_answer in enumerate(correct_answers):
                correct_parts.append(f"答案{i + 1} = {correct_answer:.2f}")
            correct_answer_message += ", ".join(correct_parts)

        # 更新会话状态
        new_problem_data = None
        response_data = {}

        if all_correct:
            # 回答正确后更新用户完成状态
            completed_all = update_user_completion_status(user_id)

            session['current_problem']['answered_correctly'] = True
            next_problem_id = problem_id + 1 if problem_id < total_problems else None

            # 构建成功消息
            message = f'🎉 回答正确！用时 {time_taken:.1f}秒，尝试 {total_attempts} 次。'

            response_data.update({
                'correct': True,
                'message': message,
                'time_taken': time_taken,
                'total_attempts': total_attempts,
                'next_problem': next_problem_id,
                'completed_all': completed_all if problem_id >= total_problems else False
            })

            # 如果完成所有题目，生成验证链接
            if completed_all and problem_id >= total_problems:
                verification_url = f"/api/user/{user_id}/completion"
                session['verification_url'] = verification_url
                response_data['verification_url'] = verification_url
                response_data['completion_message'] = '恭喜您完成了所有题目！'
        else:
            session['current_problem']['answered_correctly'] = False
            next_problem_id = problem_id

            # 答错时生成新题目（不再限制尝试次数）
            new_token, new_problem_data = fetch_problem_from_pool(actual_id)
            if not new_problem_data:
                new_token, new_problem_data = generate_and_cache_problem(actual_id)

            if new_problem_data:
                # 更新session中的题目数据和正确答案
                session['current_problem']['token'] = new_token
                session['current_problem']['start_time'] = time.time()

                # 构建错误消息
                message = f'❌ 答案不正确！用时 {time_taken:.1f}秒。{correct_answer_message}。已为您生成新题目，请重新作答。'

                response_data.update({
                    'correct': False,
                    'message': message,
                    'correct_answers': correct_answers,
                    'total_attempts': total_attempts,
                    'next_problem': next_problem_id,
                    'new_problem_generated': True,
                    'new_var_values': new_problem_data['var_values'],
                    'new_correct_answers': new_problem_data['correct_answers'],
                    'new_problem_text': new_problem_data['problem_text'],
                    'token': new_token
                })

                logger.info("已生成新题目: problem_id=%s actual_id=%s", problem_id, actual_id)
                logger.debug(
                    "新题目详情: var_values=%s correct_answers=%s",
                    new_problem_data['var_values'],
                    new_problem_data['correct_answers']
                )
            else:
                # 题目生成失败
                message = f'❌ 答案不正确！{correct_answer_message}。题目刷新失败，请重试。'
                response_data.update({
                    'correct': False,
                    'message': message,
                    'correct_answers': correct_answers,
                    'total_attempts': total_attempts,
                    'next_problem': next_problem_id,
                    'new_problem_generated': False
                })

        # 确保session被修改
        session.modified = True

        # 基础响应数据
        response_data.update({
            'success': True,
            'save_success': save_success,
            'user_answers': user_answers,
            'answer_count': answer_count,
            'problem_id': problem_id,
            'actual_id': actual_id
        })

        logger.info(
            "提交处理完成: user_id=%s problem_id=%s actual_id=%s correct=%s total_attempts=%s",
            user_id,
            problem_id,
            actual_id,
            response_data.get('correct'),
            response_data.get('total_attempts'),
        )
        logger.debug("返回数据详情: %s", response_data)
        return jsonify(response_data)

    except ValueError as e:
        logger.error("数值转换错误: %s", e)
        return jsonify({'success': False, 'message': '请输入有效的数字格式'})
    except KeyError as e:
        logger.error("缺少必要字段: %s", e)
        return jsonify({'success': False, 'message': f'缺少必要字段: {str(e)}'})
    except Exception as e:
        logger.exception("服务器错误: %s", e)
        return jsonify({'success': False, 'message': f'服务器错误: {str(e)}'})


@app.route('/admin')
@login_required
def admin_dashboard():
    """管理员主页"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    selected_paper_id = resolve_selected_exam_paper_id(
        request.args.get('paper_id', type=int),
        include_disabled_for_admin=True
    )
    selected_paper = get_exam_paper_by_id(selected_paper_id) if selected_paper_id else None
    stats = get_completion_stats(selected_paper_id)
    total_problems = get_total_problem_count(selected_paper_id)
    recent_completions = get_students_by_completion(completed=True, paper_id=selected_paper_id)
    incomplete_students = get_students_by_completion(completed=False, paper_id=selected_paper_id)

    return render_template('admin_dashboard.html',
                           stats=stats,
                           recent_completions=recent_completions,
                           incomplete_students=incomplete_students,
                           total_problems=total_problems,
                           exam_papers=get_exam_papers(),
                           selected_paper=selected_paper,
                           selected_paper_id=selected_paper_id)


@app.route('/admin/import/students', methods=['POST'])
@login_required
def admin_import_students():
    """管理员批量导入学生（xlsx：第1列学号，第2列姓名，第3列专业，第4列班级）"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    upload_file = request.files.get('students_file')
    if not upload_file or upload_file.filename == '':
        flash('请先选择要导入的 xlsx 文件', 'danger')
        return redirect(url_for('admin_dashboard'))

    if not allowed_excel_file(upload_file.filename):
        flash('文件格式错误，仅支持 .xlsx', 'danger')
        return redirect(url_for('admin_dashboard'))

    try:
        workbook = load_workbook(upload_file, read_only=True, data_only=True)
        sheet = workbook.active

        raw_rows = []
        for row in sheet.iter_rows(values_only=True):
            student_id = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
            student_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            major = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            class_name = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
            if not student_id and not student_name and not major and not class_name:
                continue
            raw_rows.append((student_id, student_name, major, class_name))
    except Exception as e:
        flash(f'读取 xlsx 失败：{e}', 'danger')
        return redirect(url_for('admin_dashboard'))

    if not raw_rows:
        flash('未读取到有效数据，请检查文件内容', 'warning')
        return redirect(url_for('admin_dashboard'))

    normalized_rows = raw_rows
    first_id, first_name, first_major, first_class = raw_rows[0]
    if first_id.lower() in {'学号', 'student_id', 'studentid', 'id', '账号', '用户名'}:
        if (first_name.lower() in {'姓名', 'name', '学生姓名'} or not first_name) and \
                (first_major.lower() in {'专业', 'major', 'major_name'} or not first_major) and \
                (first_class.lower() in {'班级', 'class', 'class_name'} or not first_class):
            normalized_rows = raw_rows[1:]

    valid_rows = []
    skipped_rows = 0
    for student_id, student_name, major, class_name in normalized_rows:
        if not student_id:
            skipped_rows += 1
            continue
        valid_rows.append((student_id, student_name or None, major or None, class_name or None))

    if not valid_rows:
        flash('未找到可导入的学号数据', 'warning')
        return redirect(url_for('admin_dashboard'))

    conn = get_db_connection()
    if not conn:
        flash('数据库连接失败', 'danger')
        return redirect(url_for('admin_dashboard'))

    inserted_count = 0
    updated_count = 0
    cursor = None
    try:
        cursor = conn.cursor()
        default_password_hash = generate_password_hash(DEFAULT_PASSWORD)
        for student_id, student_name, major, class_name in valid_rows:
            cursor.execute("SELECT id FROM users WHERE username = %s", (student_id,))
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    """
                    UPDATE users
                    SET name = %s,
                        major = %s,
                        class_name = %s,
                        password = %s,
                        password_changed = FALSE
                    WHERE username = %s
                    """,
                    (student_name, major, class_name, default_password_hash, student_id)
                )
                updated_count += 1
            else:
                cursor.execute(
                    """
                    INSERT INTO users (username, name, major, class_name, password, password_changed, avatar_filename)
                    VALUES (%s, %s, %s, %s, %s, FALSE, %s)
                    """,
                    (student_id, student_name, major, class_name, default_password_hash, DEFAULT_AVATAR)
                )
                inserted_count += 1

        conn.commit()
        flash(
            f'导入完成：新增 {inserted_count} 人，更新 {updated_count} 人，跳过 {skipped_rows} 行。'
            f'初始密码统一为 {DEFAULT_PASSWORD}。',
            'success'
        )
    except Exception as e:
        conn.rollback()
        flash(f'导入失败：{e}', 'danger')
    finally:
        if cursor:
            cursor.close()
        conn.close()

    return redirect(url_for('admin_dashboard'))


@app.route('/admin/export/students/<status>')
@login_required
def admin_export_students(status):
    """导出已完成/未完成学生列表"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    if status not in {'completed', 'incomplete'}:
        flash('无效的导出类型', 'danger')
        return redirect(url_for('admin_dashboard'))

    completed = (status == 'completed')
    selected_paper_id = resolve_selected_exam_paper_id(request.args.get('paper_id', type=int))
    students = get_students_by_completion(completed=completed, paper_id=selected_paper_id)
    total_problems = get_total_problem_count(selected_paper_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        '学生ID',
        '账号',
        '姓名',
        '专业',
        '班级',
        '完成状态',
        '完成题目数',
        '总题目数',
        '完成时间',
        '总用时(秒)',
        '注册时间'
    ])

    status_label = '已完成' if completed else '未完成'
    for student in students:
        completed_at = student['completed_at'].strftime('%Y-%m-%d %H:%M:%S') if student['completed_at'] else ''
        created_at = student['created_at'].strftime('%Y-%m-%d %H:%M:%S') if student['created_at'] else ''
        writer.writerow([
            student['id'],
            student['username'],
            student.get('name') or '',
            student.get('major') or '',
            student.get('class_name') or '',
            status_label,
            student['total_score'] or 0,
            total_problems,
            completed_at,
            round(student['total_time'] or 0, 1),
            created_at
        ])

    filename = f"students_{status}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response = Response(output.getvalue(), mimetype='text/csv; charset=utf-8')
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@app.route('/admin/exam_papers', methods=['POST'])
@login_required
def admin_create_exam_paper():
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))
    name = (request.form.get('name') or '').strip()
    description = (request.form.get('description') or '').strip()
    is_enabled = request.form.get('is_enabled') == '1'
    if not name:
        flash('题库名称不能为空', 'danger')
        return redirect(url_for('admin_dashboard'))
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO exam_papers (name, description, is_enabled) VALUES (%s, %s, %s)", (name, description or None, is_enabled))
        conn.commit()
        flash('题库创建成功', 'success')
    except mysql.connector.Error as err:
        conn.rollback()
        flash(f'题库创建失败：{err}', 'danger')
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/exam_papers/<int:paper_id>/toggle', methods=['POST'])
@login_required
def admin_toggle_exam_paper(paper_id):
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, name, is_enabled FROM exam_papers WHERE id = %s", (paper_id,))
        paper = cursor.fetchone()
        if not paper:
            flash('题库不存在', 'danger')
            return redirect(url_for('admin_dashboard'))
        new_status = not bool(paper['is_enabled'])
        cursor.execute("UPDATE exam_papers SET is_enabled = %s WHERE id = %s", (new_status, paper_id))
        conn.commit()
        flash(f"题库《{paper['name']}》已{'开启' if new_status else '关闭'}", 'success')
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('admin_dashboard', paper_id=paper_id))


@app.route('/admin/add_problem', methods=['GET', 'POST'])
@login_required
def admin_add_problem():
    """添加新题目"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        try:
            template_name = request.form['template_name']
            problem_text = request.form['problem_text']
            variables = request.form['variables']
            solution_formula = request.form['solution_formula']
            answer_count = int(request.form.get('answer_count', 1))
            difficulty = request.form.get('difficulty', 'medium')
            answer_units = request.form.get('answer_units', '')
            paper_id = request.form.get('paper_id', type=int)

            # 处理图片上传
            image_filename = None
            if 'problem_image' in request.files:
                file = request.files['problem_image']
                if file and file.filename != '':
                    image_filename = save_uploaded_file(file)
                    if image_filename:
                        # 在题目内容中插入图片
                        img_html = f'<div class="text-center mb-3"><img src="/static/images/{image_filename}" alt="{template_name}" class="problem-image img-fluid"><div class="image-caption text-muted">图：{template_name}</div></div>'
                        problem_text = img_html + problem_text

            conn = get_db_connection()
            cursor = conn.cursor()

            # 插入新题目模板
            cursor.execute("""
                INSERT INTO problem_templates 
                (template_name, problem_text, variables, solution_formula, answer_count, answer_units, difficulty, image_filename, paper_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (template_name, problem_text, variables, solution_formula, answer_count, answer_units, difficulty, image_filename, paper_id))

            conn.commit()
            cursor.close()
            conn.close()

            flash('题目添加成功！', 'success')
            return redirect(url_for('admin_manage_problems'))

        except Exception as e:
            print(f"添加题目失败: {str(e)}")
            flash(f'添加题目失败: {str(e)}', 'danger')

    return render_template('admin_add_problem.html', exam_papers=get_exam_papers())


@app.route('/admin/manage_problems')
@login_required
def admin_manage_problems():
    """管理所有题目"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    selected_paper_id = request.args.get('paper_id', type=int)
    exam_papers = get_exam_papers()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        if selected_paper_id:
            cursor.execute("SELECT * FROM problem_templates WHERE paper_id = %s ORDER BY id", (selected_paper_id,))
        else:
            cursor.execute("SELECT * FROM problem_templates ORDER BY id")
        templates = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

    display_mapping = get_problem_display_info(selected_paper_id, enabled_only=False)
    for template in templates:
        template['display_number'] = get_display_number(template['id'], template.get('paper_id'))

    return render_template('admin_manage_problems.html',
                           templates=templates,
                           display_mapping=display_mapping,
                           exam_papers=exam_papers,
                           selected_paper_id=selected_paper_id)


@app.route('/admin/edit_problem/<int:template_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_problem(template_id):
    """编辑题目"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        try:
            template_name = request.form['template_name']
            problem_text = request.form['problem_text']
            variables = request.form['variables']
            solution_formula = request.form['solution_formula']
            answer_count = int(request.form.get('answer_count', 1))
            difficulty = request.form.get('difficulty', 'medium')
            answer_units = request.form.get('answer_units', '')
            paper_id = request.form.get('paper_id', type=int)
            remove_image = request.form.get('remove_image') == 'true'
            current_image = request.form.get('current_image', '')

            # 处理图片更新
            image_filename = current_image if current_image else None

            if remove_image:
                # 删除图片
                image_filename = None
                # 从问题文本中移除图片
                problem_text = re.sub(r'<div class="text-center mb-3">.*?</div>', '', problem_text, count=1,
                                      flags=re.DOTALL)
            elif 'problem_image' in request.files:
                file = request.files['problem_image']
                if file and file.filename != '':
                    # 上传新图片
                    image_filename = save_uploaded_file(file)
                    if image_filename:
                        # 在题目内容中插入图片
                        img_html = f'<div class="text-center mb-3"><img src="/static/images/{image_filename}" alt="{template_name}" class="problem-image img-fluid"><div class="image-caption text-muted">图：{template_name}</div></div>'
                        # 移除旧的图片并添加新的
                        problem_text = re.sub(r'<div class="text-center mb-3">.*?</div>', '', problem_text, count=1,
                                              flags=re.DOTALL)
                        problem_text = img_html + problem_text

            cursor.execute("""
                UPDATE problem_templates 
                SET template_name = %s, problem_text = %s, variables = %s, 
                    solution_formula = %s, answer_count = %s, answer_units = %s, difficulty = %s, image_filename = %s, paper_id = %s
                WHERE id = %s
            """, (template_name, problem_text, variables, solution_formula, answer_count, answer_units, difficulty, image_filename, paper_id,
                  template_id))

            conn.commit()
            flash('题目更新成功！', 'success')
            return redirect(url_for('admin_manage_problems'))

        except Exception as e:
            print(f"更新题目失败: {str(e)}")
            flash(f'更新题目失败: {str(e)}', 'danger')

    # 获取题目信息
    cursor.execute("SELECT * FROM problem_templates WHERE id = %s", (template_id,))
    template = cursor.fetchone()
    cursor.close()
    conn.close()

    if not template:
        flash('题目不存在', 'danger')
        return redirect(url_for('admin_manage_problems'))

    return render_template('admin_edit_problem.html', template=template, exam_papers=get_exam_papers())


@app.route('/admin/delete_problem/<int:template_id>')
@login_required
def admin_delete_problem(template_id):
    """删除题目并清理相关图片"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 先获取题目的图片信息
        cursor.execute("SELECT image_filename FROM problem_templates WHERE id = %s", (template_id,))
        template = cursor.fetchone()

        # 删除相关的答题记录
        cursor.execute("DELETE FROM user_responses WHERE template_id = %s", (template_id,))
        # 删除题目模板
        cursor.execute("DELETE FROM problem_templates WHERE id = %s", (template_id,))

        conn.commit()

        # 删除题目后刷新所有用户的完成状态和统计
        try:
            updated_count = update_all_users_completion_status()
            print(f"ℹ️ 已刷新 {updated_count} 个用户的完成状态")
        except Exception as update_error:
            print(f"⚠️ 删除题目后刷新完成状态失败: {update_error}")

        # 如果题目有专属图片，检查并删除图片文件
        if template and template['image_filename']:
            try:
                image_path = os.path.join(app.config['UPLOAD_FOLDER'], template['image_filename'])
                if os.path.exists(image_path):
                    # 检查是否还有其他题目使用这个图片
                    cursor.execute("SELECT COUNT(*) as usage_count FROM problem_templates WHERE image_filename = %s",
                                   (template['image_filename'],))
                    usage = cursor.fetchone()
                    if usage['usage_count'] == 0:
                        os.remove(image_path)
                        print(f"✅ 已删除未使用的图片: {template['image_filename']}")
                    else:
                        print(f"ℹ️ 图片 {template['image_filename']} 仍被其他题目使用，保留文件")
            except Exception as e:
                print(f"⚠️ 删除图片文件失败: {e}")

        flash('题目删除成功！', 'success')
    except Exception as e:
        print(f"❌ 删除题目失败: {str(e)}")
        flash(f'删除题目失败: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('admin_manage_problems'))


@app.route('/admin/update_all_status')
@login_required
def update_all_status():
    """批量更新所有用户的完成状态"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        updated_count = update_all_users_completion_status()
        flash(f'成功更新 {updated_count} 个用户的完成状态', 'success')

    except Exception as e:
        print(f"批量更新状态失败: {e}")
        flash('更新状态失败', 'danger')
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('admin_dashboard'))


# 用户完成状态API
@app.route('/api/user/<int:user_id>/completion')
def check_completion(user_id):
    """检查用户完成状态 - 公开API"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT 
            u.username,
            u.completed_all,
            u.completed_at,
            u.total_score,
            u.total_time,
            COUNT(DISTINCT r.template_id) as completed_problems,
            MAX(r.response_time) as last_completion_time
        FROM users u
        LEFT JOIN user_responses r ON u.id = r.user_id AND r.is_correct = TRUE
        WHERE u.id = %s
        GROUP BY u.id
    """, (user_id,))

    result = cursor.fetchone()
    cursor.close()
    conn.close()

    if result:
        return jsonify({
            'user_id': user_id,
            'username': result['username'],
            'completed_all': result['completed_all'],
            'completed_problems': result['completed_problems'],
            'total_score': result['total_score'],
            'total_time': result['total_time'],
            'completed_at': result['completed_at'],
            'last_completion_time': result['last_completion_time'],
            'verified': True
        })
    else:
        return jsonify({
            'user_id': user_id,
            'error': '用户不存在',
            'verified': False
        }), 404


# 管理员功能 - 学生完成情况
@app.route('/admin/students/<status>')
@login_required
def admin_students_by_status(status):
    """按完成状态查看学生列表"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    selected_paper_id = request.args.get('paper_id', type=int)
    completed = (status == 'completed')
    students = get_students_by_completion(completed=completed, paper_id=selected_paper_id)
    total_problems = get_total_problem_count(selected_paper_id)
    completion_stats = get_completion_stats(selected_paper_id)
    class_stats = get_class_comparison_stats(selected_paper_id)
    total_students = completion_stats['stats']['total_students'] or 0
    completed_students = completion_stats['stats']['completed_count'] or 0

    status_text = '已完成' if completed else '未完成'

    return render_template('admin_students.html',
                           students=students,
                           status=status,
                           status_text=status_text,
                           total_problems=total_problems,
                           total_students=total_students,
                           completed_students=completed_students,
                           class_stats=class_stats,
                           username=session['username'])


# 管理员功能 - 学生详细答题情况
def infer_knowledge_label(template_name):
    """根据题目名称推断知识点标签。"""
    name = (template_name or '').strip()
    rules = [
        ('电磁学', ['电磁', '磁场', '磁铁', '感应', '线圈', '电流', '电压', '电阻', '电荷', '安培', '法拉第']),
        ('力学', ['力学', '受力', '牛顿', '加速度', '速度', '位移', '动量', '机械', '弹簧', '摩擦', '圆周', '功', '能量']),
        ('热学', ['热', '温度', '内能', '热量', '压强', '气体']),
        ('光学', ['光', '透镜', '折射', '反射', '干涉', '衍射']),
        ('振动与波', ['波', '振动', '频率', '波长', '声', '共振']),
    ]
    for label, keywords in rules:
        if any(keyword in name for keyword in keywords):
            return label
    return '综合分析'


def build_student_insight_summary(problem_stats, knowledge_stats, error_type_stats):
    """生成学生表现自动结论。"""
    solved = [stat for stat in problem_stats if stat.get('is_completed')]
    total = len(problem_stats)
    solved_count = len(solved)

    summary = []
    if total:
        summary.append(f'共完成 {solved_count}/{total} 道题，整体完成率 {round(solved_count / total * 100, 1)}%。')

    comparable = [item for item in knowledge_stats if item.get('attempted_templates')]
    if comparable:
        best = max(comparable, key=lambda item: (item.get('correct_rate', 0), item.get('avg_attempts', 999) * -1))
        weakest = min(comparable, key=lambda item: (item.get('correct_rate', 0), -(item.get('avg_attempts', 0))))
        if best.get('label') == weakest.get('label'):
            summary.append(f"当前题库主要集中在{best['label']}，该知识点正确率为 {best.get('correct_rate', 0):.1f}%。")
        else:
            summary.append(f"你在 {best['label']} 题上表现最好，正确率 {best.get('correct_rate', 0):.1f}%；{weakest['label']} 仍是当前薄弱点，正确率 {weakest.get('correct_rate', 0):.1f}%。")

    top_error = next((item for item in error_type_stats if item.get('error_type') and item.get('error_type') != '正确'), None)
    if top_error and top_error.get('count'):
        summary.append(f"最近错题主要集中在“{top_error['error_type']}”类型，共出现 {top_error['count']} 次，建议优先复盘对应步骤。")

    if not summary:
        summary.append('当前作答数据较少，继续完成更多题目后可生成更稳定的学习结论。')

    return ' '.join(summary)

@app.route('/admin/student/<int:user_id>/details')
@login_required
def admin_student_details(user_id):
    """查看学生详细答题情况"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    selected_paper_id = resolve_selected_exam_paper_id(request.args.get('paper_id', type=int))
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 获取学生基本信息
        cursor.execute(
            "SELECT username, name, class_name, completed_all, total_score, total_time FROM users WHERE id = %s",
            (user_id,)
        )
        student = cursor.fetchone()

        if not student:
            flash('学生不存在', 'danger')
            return redirect(url_for('admin_dashboard'))

        # 获取题目总数
        total_problems = get_total_problem_count(selected_paper_id)
        problem_template_filter, problem_template_params = build_enabled_paper_filter('t', selected_paper_id)
        response_filter, response_params = build_enabled_paper_filter('ur', selected_paper_id)

        # 每题作答状态：是否答对；答对时展示达到答对所需次数与累计时长
        cursor.execute(f"""
            WITH attempt_summary AS (
                SELECT
                    template_id,
                    attempt_count,
                    COUNT(*) AS total_answers,
                    SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) AS correct_answers,
                    MAX(time_taken) AS time_taken,
                    MAX(response_time) AS last_response_time,
                    CASE WHEN SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*) THEN 1 ELSE 0 END AS is_fully_correct
                FROM user_responses ur
                WHERE user_id = %s {response_filter}
                GROUP BY template_id, attempt_count
            ),
            progress AS (
                SELECT
                    template_id,
                    COUNT(*) AS total_attempts,
                    SUM(is_fully_correct) AS correct_attempts,
                    SUM(time_taken) AS total_time_spent,
                    MAX(last_response_time) AS last_attempt_time,
                    MAX(is_fully_correct) AS is_completed,
                    MIN(CASE WHEN is_fully_correct = 1 THEN attempt_count ELSE NULL END) AS attempts_to_correct
                FROM attempt_summary
                GROUP BY template_id
            )
            SELECT
                t.id AS template_id,
                t.template_name,
                COALESCE(p.total_attempts, 0) AS total_attempts,
                COALESCE(p.correct_attempts, 0) AS correct_attempts,
                COALESCE(p.total_time_spent, 0) AS total_time_spent,
                p.last_attempt_time,
                COALESCE(p.is_completed, 0) AS is_completed,
                p.attempts_to_correct,
                (
                    SELECT SUM(a2.time_taken)
                    FROM attempt_summary a2
                    WHERE a2.template_id = t.id
                      AND p.attempts_to_correct IS NOT NULL
                      AND a2.attempt_count <= p.attempts_to_correct
                ) AS cumulative_time_to_correct
            FROM problem_templates t
            LEFT JOIN progress p ON t.id = p.template_id
            WHERE 1 = 1 {problem_template_filter}
            ORDER BY t.id
        """, [user_id] + response_params + problem_template_params)

        problem_stats = cursor.fetchall()
        for stat in problem_stats:
            stat['knowledge_label'] = infer_knowledge_label(stat.get('template_name'))
            stat['correct_rate'] = 100.0 if stat.get('is_completed') else 0.0
        completed_problems_count = sum(1 for stat in problem_stats if stat.get('is_completed'))

        # 计算总体统计
        cursor.execute(f"""
            WITH attempt_summary AS (
                SELECT
                    template_id,
                    attempt_count,
                    COUNT(*) AS total_answers,
                    SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) AS correct_answers,
                    MAX(time_taken) AS time_taken
                FROM user_responses ur
                WHERE user_id = %s {response_filter}
                GROUP BY template_id, attempt_count
            )
            SELECT
                COUNT(*) as total_attempts,
                SUM(CASE WHEN correct_answers = total_answers THEN 1 ELSE 0 END) as total_correct,
                AVG(time_taken) as overall_avg_time
            FROM attempt_summary
        """, [user_id] + response_params)

        overall_stats = cursor.fetchone() or {}
        overall_stats.setdefault('total_attempts', 0)
        overall_stats.setdefault('total_correct', 0)
        overall_stats.setdefault('overall_avg_time', 0)
        if isinstance(overall_stats.get('overall_avg_time'), Decimal):
            overall_stats['overall_avg_time'] = float(overall_stats['overall_avg_time'])

        trend_points = []
        cumulative_correct = 0
        for index, stat in enumerate(problem_stats, start=1):
            cumulative_correct += 1 if stat.get('is_completed') else 0
            trend_points.append({
                'label': f"第{get_display_number(stat.get('template_id'), selected_paper_id)}题",
                'short_label': str(get_display_number(stat.get('template_id'), selected_paper_id)),
                'correct_rate': round((cumulative_correct / index) * 100, 1),
                'attempts': stat.get('total_attempts') or 0,
                'completed': bool(stat.get('is_completed'))
            })

        knowledge_map = {}
        for stat in problem_stats:
            label = stat['knowledge_label']
            bucket = knowledge_map.setdefault(label, {
                'label': label,
                'total_templates': 0,
                'attempted_templates': 0,
                'completed_templates': 0,
                'total_attempts': 0,
                'sum_attempts_to_correct': 0,
                'attempts_to_correct_count': 0,
            })
            bucket['total_templates'] += 1
            if (stat.get('total_attempts') or 0) > 0:
                bucket['attempted_templates'] += 1
            if stat.get('is_completed'):
                bucket['completed_templates'] += 1
            bucket['total_attempts'] += stat.get('total_attempts') or 0
            if stat.get('attempts_to_correct'):
                bucket['sum_attempts_to_correct'] += stat['attempts_to_correct']
                bucket['attempts_to_correct_count'] += 1

        knowledge_stats = []
        for item in knowledge_map.values():
            attempted = item['attempted_templates']
            completed = item['completed_templates']
            item['correct_rate'] = round((completed / attempted) * 100, 1) if attempted else 0
            item['avg_attempts'] = round(item['sum_attempts_to_correct'] / item['attempts_to_correct_count'], 1) if item['attempts_to_correct_count'] else None
            knowledge_stats.append(item)
        knowledge_stats.sort(key=lambda item: (-item['correct_rate'], -item['completed_templates'], item['label']))

        cursor.execute(f"""
            SELECT COALESCE(NULLIF(TRIM(error_type), ''), '未知') AS error_type, COUNT(*) AS count
            FROM user_responses ur
            WHERE user_id = %s AND is_correct = FALSE {response_filter}
            GROUP BY COALESCE(NULLIF(TRIM(error_type), ''), '未知')
            ORDER BY count DESC, error_type ASC
        """, [user_id] + response_params)
        error_type_stats = cursor.fetchall()

        cursor.execute(f"""
            SELECT
                t.id AS template_id,
                t.template_name,
                COUNT(*) AS wrong_count,
                MAX(ur.response_time) AS last_wrong_time,
                SUM(CASE WHEN ur.error_type = '计算误差' THEN 1 ELSE 0 END) AS calc_error_count,
                SUM(CASE WHEN ur.error_type = '单位错误' THEN 1 ELSE 0 END) AS unit_error_count,
                SUM(CASE WHEN ur.error_type = '格式错误' THEN 1 ELSE 0 END) AS format_error_count
            FROM user_responses ur
            JOIN problem_templates t ON t.id = ur.template_id
            WHERE ur.user_id = %s AND ur.is_correct = FALSE {response_filter}
            GROUP BY t.id, t.template_name
            ORDER BY wrong_count DESC, t.id ASC
            LIMIT 6
        """, [user_id] + response_params)
        wrong_problem_stats = cursor.fetchall()
        for item in wrong_problem_stats:
            item['knowledge_label'] = infer_knowledge_label(item.get('template_name'))

        insight_summary = build_student_insight_summary(problem_stats, knowledge_stats, error_type_stats)

        return render_template('admin_student_details.html',
                               student=student,
                               problem_stats=problem_stats,
                               overall_stats=overall_stats,
                               completed_problems_count=completed_problems_count,
                               total_problems=total_problems,
                               trend_points=trend_points,
                               knowledge_stats=knowledge_stats,
                               error_type_stats=error_type_stats,
                               wrong_problem_stats=wrong_problem_stats,
                               insight_summary=insight_summary,
                               username=session['username'],
                               get_display_number=get_display_number,
                               selected_paper_id=selected_paper_id)  # 传递函数到模板
    except Exception as e:
        print(f"获取学生详情失败: {str(e)}")
        import traceback
        print(f"详细错误: {traceback.format_exc()}")
        flash('获取学生详情失败', 'danger')
        return redirect(url_for('admin_dashboard'))
    finally:
        cursor.close()
        conn.close()


# 管理员功能 - 所有学生题目答题情况
@app.route('/admin/all_problems_stats')
@login_required
def admin_all_problems_stats():
    """查看所有学生对每道题的答题情况"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        selected_paper_id = resolve_selected_exam_paper_id(request.args.get('paper_id', type=int))
        selected_class_name = (request.args.get('class_name') or '').strip()
        selected_major = (request.args.get('major') or '').strip()
        problem_filter, problem_filter_params = build_enabled_paper_filter('t', selected_paper_id)
        response_filter, response_params = build_enabled_paper_filter('ur', selected_paper_id)
        response_filter_ur2, response_params_ur2 = build_enabled_paper_filter('ur2', selected_paper_id)

        # 加载班级筛选选项
        cursor.execute("""
            SELECT DISTINCT COALESCE(NULLIF(TRIM(class_name), ''), '未分班') AS class_name
            FROM users
            WHERE username != 'admin'
            ORDER BY class_name
        """)
        class_options = [row['class_name'] for row in cursor.fetchall()]

        cursor.execute("""
            SELECT DISTINCT COALESCE(NULLIF(TRIM(major), ''), '未设置专业') AS major
            FROM users
            WHERE username != 'admin'
            ORDER BY major
        """)
        major_options = [row['major'] for row in cursor.fetchall()]

        filters = ["u.username != 'admin'"]
        params = []

        if selected_class_name:
            if selected_class_name == '未分班':
                filters.append("(u.class_name IS NULL OR TRIM(u.class_name) = '')")
            else:
                filters.append("u.class_name = %s")
                params.append(selected_class_name)

        if selected_major:
            if selected_major == '未设置专业':
                filters.append("(u.major IS NULL OR TRIM(u.major) = '')")
            else:
                filters.append("u.major = %s")
                params.append(selected_major)

        filter_clause = ' AND '.join(filters)

        # 题目层级统计（首次正确率、总答题次数、最终答对人数、平均正确时长）
        cursor.execute(f"""
            WITH attempt_summary AS (
                SELECT
                    ur.user_id,
                    ur.template_id,
                    ur.attempt_count,
                    COUNT(*) AS total_answers,
                    SUM(CASE WHEN ur.is_correct THEN 1 ELSE 0 END) AS correct_answers,
                    MAX(ur.time_taken) AS time_taken,
                    CASE WHEN SUM(CASE WHEN ur.is_correct THEN 1 ELSE 0 END) = COUNT(*) THEN 1 ELSE 0 END AS is_fully_correct
                FROM user_responses ur
                JOIN users u ON ur.user_id = u.id
                WHERE {filter_clause} {response_filter}
                GROUP BY ur.user_id, ur.template_id, ur.attempt_count
            ),
            user_problem_stats AS (
                SELECT
                    user_id,
                    template_id,
                    COUNT(*) AS attempt_count,
                    MIN(attempt_count) AS first_attempt_no,
                    MIN(CASE WHEN is_fully_correct = 1 THEN attempt_count ELSE NULL END) AS first_correct_attempt_no,
                    MAX(is_fully_correct) AS final_is_correct
                FROM attempt_summary
                GROUP BY user_id, template_id
            )
            SELECT
                t.id as template_id,
                t.template_name,
                COUNT(DISTINCT ups.user_id) as participant_students,
                COALESCE(SUM(ups.attempt_count), 0) as total_attempts,
                SUM(CASE WHEN first_attempt.is_fully_correct = 1 THEN 1 ELSE 0 END) as first_correct_students,
                SUM(CASE WHEN ups.final_is_correct = 1 THEN 1 ELSE 0 END) as final_correct_students,
                AVG(CASE
                    WHEN ups.first_correct_attempt_no IS NOT NULL AND correct_attempt.attempt_count = ups.first_correct_attempt_no
                    THEN correct_attempt.time_taken
                    ELSE NULL
                END) as avg_correct_time
            FROM problem_templates t
            LEFT JOIN user_problem_stats ups ON t.id = ups.template_id
            LEFT JOIN attempt_summary first_attempt
                ON first_attempt.user_id = ups.user_id
               AND first_attempt.template_id = ups.template_id
               AND first_attempt.attempt_count = ups.first_attempt_no
            LEFT JOIN attempt_summary correct_attempt
                ON correct_attempt.user_id = ups.user_id
               AND correct_attempt.template_id = ups.template_id
               AND correct_attempt.attempt_count = ups.first_correct_attempt_no
            WHERE 1 = 1 {problem_filter}
            GROUP BY t.id, t.template_name
            ORDER BY t.id
        """, params + response_params + problem_filter_params)

        problem_stats_result = cursor.fetchall()
        problem_stats = []

        # 转换problem_stats中的Decimal类型，避免模板中tojson序列化失败
        for raw_stat in problem_stats_result:
            stat = dict(raw_stat)
            for key, value in stat.items():
                if isinstance(value, Decimal):
                    stat[key] = float(value)

            participants = stat.get('participant_students') or 0
            first_correct_students = stat.get('first_correct_students') or 0
            if participants > 0:
                stat['first_correct_rate'] = round((first_correct_students / participants) * 100, 1)
            else:
                stat['first_correct_rate'] = 0
            problem_stats.append(stat)

        # 获取筛选后的学生总数（排除管理员）
        student_filters = ["u.username != 'admin'"]
        student_params = []
        if selected_class_name:
            if selected_class_name == '未分班':
                student_filters.append("(u.class_name IS NULL OR TRIM(u.class_name) = '')")
            else:
                student_filters.append("u.class_name = %s")
                student_params.append(selected_class_name)
        if selected_major:
            if selected_major == '未设置专业':
                student_filters.append("(u.major IS NULL OR TRIM(u.major) = '')")
            else:
                student_filters.append("u.major = %s")
                student_params.append(selected_major)

        cursor.execute(
            f"SELECT COUNT(*) as total FROM users u WHERE {' AND '.join(student_filters)}",
            student_params
        )

        total_students_result = cursor.fetchone()
        total_students = total_students_result['total'] if total_students_result else 0

        cursor.execute(
            f"""
            WITH completed AS (
                SELECT user_id, COUNT(DISTINCT template_id) AS total_score, COALESCE(SUM(time_taken), 0) AS total_time
                FROM user_responses ur
                WHERE ur.is_correct = TRUE {response_filter}
                  AND (ur.user_id, ur.template_id, ur.attempt_count) IN (
                    SELECT user_id, template_id, attempt_count
                    FROM user_responses ur2
                    WHERE 1 = 1 {response_filter_ur2}
                    GROUP BY user_id, template_id, attempt_count
                    HAVING SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) = COUNT(*)
                  )
                GROUP BY user_id
            )
            SELECT
                u.id,
                u.username,
                u.name,
                COALESCE(NULLIF(TRIM(u.major), ''), '未设置专业') AS major,
                COALESCE(NULLIF(TRIM(u.class_name), ''), '未分班') AS class_name,
                CASE WHEN COALESCE(c.total_score, 0) >= %s AND %s > 0 THEN TRUE ELSE FALSE END AS completed_all,
                COALESCE(c.total_score, 0) AS total_score,
                COALESCE(c.total_time, 0) AS total_time,
                u.created_at,
                u.completed_at
            FROM users u
            LEFT JOIN completed c ON c.user_id = u.id
            WHERE {' AND '.join(student_filters)}
            ORDER BY u.class_name ASC, u.username ASC
            """,
            response_params + response_params_ur2 + [get_total_problem_count(selected_paper_id), get_total_problem_count(selected_paper_id)] + student_params
        )
        filtered_students = cursor.fetchall()

        # 汇总统计
        cursor.execute(f"""
            WITH attempt_summary AS (
                SELECT
                    ur.user_id,
                    ur.template_id,
                    ur.attempt_count,
                    COUNT(*) AS total_answers,
                    SUM(CASE WHEN ur.is_correct THEN 1 ELSE 0 END) AS correct_answers,
                    MAX(ur.time_taken) AS time_taken,
                    CASE WHEN SUM(CASE WHEN ur.is_correct THEN 1 ELSE 0 END) = COUNT(*) THEN 1 ELSE 0 END AS is_fully_correct
                FROM user_responses ur
                JOIN users u ON ur.user_id = u.id
                WHERE {filter_clause} {response_filter}
                GROUP BY ur.user_id, ur.template_id, ur.attempt_count
            ),
            user_problem_stats AS (
                SELECT
                    user_id,
                    template_id,
                    COUNT(*) AS attempt_count,
                    MIN(attempt_count) AS first_attempt_no,
                    MIN(CASE WHEN is_fully_correct = 1 THEN attempt_count ELSE NULL END) AS first_correct_attempt_no,
                    MAX(is_fully_correct) AS final_is_correct
                FROM attempt_summary
                GROUP BY user_id, template_id
            )
            SELECT
                COALESCE(SUM(ups.attempt_count), 0) as total_attempts,
                SUM(CASE WHEN first_attempt.is_fully_correct = 1 THEN 1 ELSE 0 END) as first_correct_students,
                SUM(CASE WHEN ups.final_is_correct = 1 THEN 1 ELSE 0 END) as final_correct_students,
                AVG(CASE
                    WHEN ups.first_correct_attempt_no IS NOT NULL AND correct_attempt.attempt_count = ups.first_correct_attempt_no
                    THEN correct_attempt.time_taken
                    ELSE NULL
                END) as avg_correct_time
            FROM user_problem_stats ups
            LEFT JOIN attempt_summary first_attempt
                ON first_attempt.user_id = ups.user_id
               AND first_attempt.template_id = ups.template_id
               AND first_attempt.attempt_count = ups.first_attempt_no
            LEFT JOIN attempt_summary correct_attempt
                ON correct_attempt.user_id = ups.user_id
               AND correct_attempt.template_id = ups.template_id
               AND correct_attempt.attempt_count = ups.first_correct_attempt_no
        """, params + response_params)
        overall_stats = cursor.fetchone() or {}
        for key, value in list(overall_stats.items()):
            if isinstance(value, Decimal):
                overall_stats[key] = float(value)
        overall_stats.setdefault('total_attempts', 0)
        overall_stats.setdefault('first_correct_students', 0)
        overall_stats.setdefault('final_correct_students', 0)
        overall_stats.setdefault('avg_correct_time', 0)

        total_participants = sum((stat.get('participant_students') or 0) for stat in problem_stats)
        total_first_correct = sum((stat.get('first_correct_students') or 0) for stat in problem_stats)
        overall_stats['first_correct_rate'] = round((total_first_correct / total_participants) * 100, 1) if total_participants else 0

        return render_template('admin_all_problems_stats.html',
                               problem_stats=problem_stats,
                               total_students=total_students,
                               overall_stats=overall_stats,
                               filtered_students=filtered_students,
                               class_options=class_options,
                               major_options=major_options,
                               selected_class_name=selected_class_name,
                               selected_major=selected_major,
                               username=session['username'],
                               selected_paper_id=selected_paper_id)

    except Exception as e:
        print(f"获取题目统计失败: {str(e)}")
        import traceback
        print(f"详细错误: {traceback.format_exc()}")
        flash(f'获取题目统计失败: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))
    finally:
        cursor.close()
        conn.close()


def is_mobile_device():
    """检测是否为移动设备"""
    user_agent = request.headers.get('User-Agent', '').lower()
    mobile_pattern = re.compile(r'mobile|android|webos|iphone|ipad|ipod|blackberry|windows phone')
    return bool(mobile_pattern.search(user_agent))


def is_touch_device():
    """检测是否为触摸设备（简化版）"""
    user_agent = request.headers.get('User-Agent', '').lower()
    touch_pattern = re.compile(r'mobile|android|iphone|ipad|ipod')
    return bool(touch_pattern.search(user_agent))


@app.context_processor
def inject_device_status():
    """向所有模板注入设备状态"""
    return {
        'is_mobile': is_mobile_device(),
        'is_touch': is_touch_device(),
        'display_name': session.get('display_name'),
        'avatar_filename': session.get('avatar_filename', DEFAULT_AVATAR),
        'is_admin': session.get('username') == 'admin' if 'username' in session else False,
        'available_avatars': get_avatar_choices(),
        'show_password_modal': session.pop('show_password_modal', False)
    }


@app.route('/api/user/completion_status')
@login_required
def api_user_completion_status():
    """获取用户所有题目的完成状态"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        selected_paper_id = resolve_selected_exam_paper_id()
        templates = get_problem_templates_by_paper(selected_paper_id)

        completion_status = {}

        for template in templates:
            template_id = template['id']
            # 检查该题目是否已完成（有正确答题记录）
            completion_status[template_id] = has_full_correct_attempt(cursor, session['user_id'], template_id)

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'completion_status': completion_status
        })

    except Exception as e:
        print(f"获取完成状态失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取完成状态失败: {str(e)}'
        })


# 图片管理功能
@app.route('/admin/image_manager')
@login_required
def admin_image_manager():
    """图片管理页面"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    import os
    from datetime import datetime

    images = []
    images_path = os.path.join(app.root_path, 'static', 'images')

    if os.path.exists(images_path):
        for filename in os.listdir(images_path):
            if allowed_file(filename):
                file_path = os.path.join(images_path, filename)
                file_stat = os.stat(file_path)
                images.append({
                    'filename': filename,
                    'size': round(file_stat.st_size / 1024, 1),  # KB
                    'modified': datetime.fromtimestamp(file_stat.st_mtime).strftime('%Y-%m-%d %H:%M')
                })

    return render_template('admin_image_manager.html', images=images)


@app.route('/admin/delete_image/<filename>')
@login_required
def admin_delete_image(filename):
    """删除图片"""
    if session.get('username') != 'admin':
        flash('权限不足', 'danger')
        return redirect(url_for('dashboard'))

    import os
    from werkzeug.utils import secure_filename

    # 安全检查
    filename = secure_filename(filename)
    image_path = os.path.join(app.root_path, 'static', 'images', filename)

    if os.path.exists(image_path):
        try:
            # 检查是否有题目在使用这个图片
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id, template_name FROM problem_templates WHERE image_filename = %s", (filename,))
            using_templates = cursor.fetchall()
            cursor.close()
            conn.close()

            if using_templates:
                template_names = [t['template_name'] for t in using_templates]
                flash(f'无法删除图片，以下题目正在使用：{", ".join(template_names)}', 'danger')
            else:
                os.remove(image_path)
                flash('图片删除成功！', 'success')
        except Exception as e:
            print(f"删除图片失败: {str(e)}")
            flash(f'删除图片失败: {str(e)}', 'danger')
    else:
        flash('图片不存在', 'danger')

    return redirect(url_for('admin_image_manager'))


@app.route('/health')
def health_check():
    """健康检查端点"""
    try:
        # 检查数据库连接
        conn = get_db_connection()
        if conn and conn.is_connected():
            conn.close()
            return jsonify({'status': 'healthy', 'database': 'connected'})
        else:
            return jsonify({'status': 'unhealthy', 'database': 'disconnected'}), 500
    except Exception as e:
        return jsonify({'status': 'unhealthy', 'error': str(e)}), 500


@app.route('/api/exam/status')
@login_required
def exam_system_status():
    """考试系统状态监控"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # 获取系统统计
    cursor.execute("SELECT COUNT(*) as active_users FROM users WHERE completed_all = FALSE")
    active_users = cursor.fetchone()

    cursor.execute("SELECT COUNT(*) as total_attempts FROM user_responses WHERE DATE(response_time) = CURDATE()")
    today_attempts = cursor.fetchone()

    cursor.close()
    conn.close()

    return jsonify({
        'active_users': active_users['active_users'],
        'today_attempts': today_attempts['total_attempts'],
        'server_time': datetime.now().isoformat(),
        'status': 'operational'
    })


if __name__ == '__main__':
    # 确保图片目录存在
    os.makedirs('static/images', exist_ok=True)


    # 初始化数据库
    initialize_database()

    # 确保用户表字段完整
    ensure_user_columns()

    # 确保默认图片存在
    ensure_default_images()

    # 创建管理员用户
    create_admin_user()

    # 修复可能存在的表结构问题
    repair_database()

    # 运行数据库诊断
    print("运行数据库诊断...")
    diagnose_database_issue()

    # 验证图片一致性
    print("验证图片一致性...")
    images_ok = verify_image_consistency()
    if not images_ok:
        print("⚠️ 警告: 部分图片文件缺失，请检查以上列表")

    prewarm_flag = os.getenv('PREWARM') == '1' or os.getenv('PORT') == '5000'
    if prewarm_flag:
        prewarm_pools()
    else:
        print("[PREWARM] 跳过预热，未满足 PREWARM 或端口条件")

    app.run(host='0.0.0.0', port=5000, debug=False)
