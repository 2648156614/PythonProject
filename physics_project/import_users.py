import re
from openpyxl import load_workbook
import mysql.connector

# ==================ã€ä½ åªéœ€è¦æ”¹è¿™é‡Œã€‘==================

EXCEL_PATH = "å®éªŒæ¨¡æ¿.xlsx"   # Excel æ–‡ä»¶è·¯å¾„ï¼ˆå¯ç›¸å¯¹/ç»å¯¹ï¼‰
DEFAULT_PASSWORD = "123456"

DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "123456",
    "database": "physics_new3",
}

IMPORT_MODE = "skip"
# skip   = å·²å­˜åœ¨å­¦å·ç›´æ¥è·³è¿‡
# update = å·²å­˜åœ¨å­¦å·åˆ™æ›´æ–°å§“åï¼ˆå’Œå¯†ç ï¼‰

RESET_PASSWORD = True
# True  = å¯¼å…¥æ—¶å¯†ç ç»Ÿä¸€è®¾ä¸º 123456
# False = ä¸åŠ¨å·²æœ‰ç”¨æˆ·å¯†ç 

# =====================================================

ID_HEADERS = {
    "å­¦å·", "å­¦ç”Ÿå­¦å·", "å­¦å·(å¿…å¡«)", "student_id", "studentid", "id", "è´¦å·", "ç”¨æˆ·å", "username"
}
NAME_HEADERS = {"å§“å", "åå­—", "name", "student_name"}


def clean_student_id(value):
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None

    # å¤„ç† 20230001.0
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]

    # ç§‘å­¦è®¡æ•°æ³•
    if "e" in s.lower():
        try:
            s = str(int(float(s)))
        except Exception:
            return None

    s = re.sub(r"\s+", "", s)
    s = re.sub(r"\D", "", s)
    return s if s else None


def normalize(s):
    return str(s).strip().lower() if s else ""


def find_col(header, candidates):
    for i, h in enumerate(header):
        if normalize(h) in {normalize(x) for x in candidates}:
            return i
    return None


def main():
    print("ğŸš€ å¼€å§‹å¯¼å…¥ç”¨æˆ·æ•°æ®...")

    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            print("âš ï¸ Excel ä¸ºç©º")
            return

        header = rows[0]
        id_idx = find_col(header, ID_HEADERS)
        name_idx = find_col(header, NAME_HEADERS)

        if id_idx is None:
            print("âŒ æœªæ‰¾åˆ°å­¦å·åˆ—")
            print("å¯è¯†åˆ«è¡¨å¤´ï¼š", ID_HEADERS)
            return

        print(f"âœ… å­¦å·åˆ—ä½ç½®ï¼šç¬¬ {id_idx + 1} åˆ—")

        if IMPORT_MODE == "skip":
            if RESET_PASSWORD:
                sql = """
                INSERT IGNORE INTO users (username, password, name)
                VALUES (%s, %s, %s)
                """
            else:
                sql = """
                INSERT IGNORE INTO users (username, name)
                VALUES (%s, %s)
                """
        else:
            if RESET_PASSWORD:
                sql = """
                INSERT INTO users (username, password, name)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    name = VALUES(name),
                    password = VALUES(password)
                """
            else:
                sql = """
                INSERT INTO users (username, name)
                VALUES (%s, %s)
                ON DUPLICATE KEY UPDATE
                    name = VALUES(name)
                """

        batch = []
        skipped = 0

        for r in rows[1:]:
            sid = clean_student_id(r[id_idx] if id_idx < len(r) else None)
            if not sid:
                skipped += 1
                continue

            name = ""
            if name_idx is not None and name_idx < len(r):
                name = str(r[name_idx]).strip() if r[name_idx] else ""

            if RESET_PASSWORD:
                batch.append((sid, DEFAULT_PASSWORD, name))
            else:
                batch.append((sid, name))

        if not batch:
            print("âš ï¸ æ²¡æœ‰å¯å¯¼å…¥çš„æ•°æ®")
            return

        cursor.executemany(sql, batch)
        conn.commit()

        print("ğŸ‰ å¯¼å…¥å®Œæˆï¼")
        print(f"   å°è¯•å¯¼å…¥ï¼š{len(batch)} æ¡")
        print(f"   è·³è¿‡æ— æ•ˆï¼š{skipped} æ¡")
        print(f"   æ¨¡å¼ï¼š{IMPORT_MODE}")

    except Exception as e:
        conn.rollback()
        print("âŒ å¯¼å…¥å¤±è´¥ï¼š", e)
    finally:
        cursor.close()
        conn.close()


if __name__ == "__main__":
    main()
